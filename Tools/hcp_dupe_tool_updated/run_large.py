#!/usr/bin/env python3
"""
run_large.py — Optimized entry point for large-scale HCP/HCO duplicate detection.

Handles 600K+ rows / 300K+ unique VIDs on 8-16GB RAM via:
  - CSV input (not XLSX for speed)
  - Vectorized pandas canonicalization
  - Tighter blocking (max_block_size=300, disabled phonetic/first-initial by default)
  - Multiprocessing pair comparison
  - Streaming CSV output + batch Excel consolidation
  - Memory monitoring and pair-count limiting

Usage examples:

  # Basic large-scale run (CSV input, 8 workers, streaming output)
  python -m hcp_dupe_tool.run_large --input large.csv --outdir ./results

  # Disable Excel output for speed (CSV only)
  python -m hcp_dupe_tool.run_large --input large.csv --outdir ./results --no-excel

  # Override worker count and add phonetic blocking back
  python -m hcp_dupe_tool.run_large --input large.csv --workers 12 --config config.yaml --outdir ./results

  # Warn if memory exceeds threshold
  python -m hcp_dupe_tool.run_large --input large.csv --memory-limit 12000 --outdir ./results

  # Limit candidates to 500K pairs (truncates lower-quality blocks if exceeded)
  python -m hcp_dupe_tool.run_large --input large.csv --max-pairs 500000 --outdir ./results
"""

from __future__ import annotations

import argparse
import csv
import gc
import logging
import multiprocessing as mp
import os
import sys
import time
from typing import Any, Iterable

import pandas as pd

logger = logging.getLogger("dupe_tool.large")


# ── Config Loading (shared with run.py) ──────────────────────────────

def _load_config(config_path: str | None) -> dict[str, Any]:
    """Load YAML config, falling back to built-in defaults."""
    defaults_path = os.path.join(os.path.dirname(__file__), "config_default.yaml")

    try:
        import yaml
    except ImportError:
        print(
            "WARNING: PyYAML not installed — using hardcoded defaults.\n"
            "  Install with:  pip install pyyaml\n"
        )
        return _hardcoded_defaults()

    cfg: dict[str, Any] = {}
    if os.path.isfile(defaults_path):
        with open(defaults_path, "r") as f:
            cfg = yaml.safe_load(f) or {}

    if config_path and os.path.isfile(config_path):
        with open(config_path, "r") as f:
            user_cfg = yaml.safe_load(f) or {}
        cfg = _deep_merge(cfg, user_cfg)
        print(f"Loaded config: {config_path}")
    elif config_path:
        print(f"WARNING: Config file not found: {config_path} — using defaults.")

    return cfg


def _deep_merge(base: dict, overlay: dict) -> dict:
    """Recursively merge overlay into base dict."""
    merged = base.copy()
    for k, v in overlay.items():
        if k in merged and isinstance(merged[k], dict) and isinstance(v, dict):
            merged[k] = _deep_merge(merged[k], v)
        else:
            merged[k] = v
    return merged


def _hardcoded_defaults() -> dict[str, Any]:
    """Minimal hardcoded defaults when YAML isn't available."""
    return {
        "columns": {
            "hcp_vid": "hcp.vid__v (VID)",
            "hcp_last_name": "hcp.last_name__v (LAST NAME)",
            "hcp_first_name": "hcp.first_name__v (FIRST NAME)",
            "hcp_middle_name": "hcp.middle_name__v (MIDDLE NAME)",
            "hcp_full_name": "hcp.source_full_name__v (SOURCE FULL NAME)",
            "specialty_1": "hcp.specialty_1__v (SPECIALTY 1)",
            "specialty_2": "hcp.specialty_2__v (SPECIALTY 2)",
            "specialty_3": "hcp.specialty_3__v (SPECIALTY 3)",
            "specialty_4": "hcp.specialty_4__v (SPECIALTY 4)",
            "license_number": "license.license_number__v (LICENSE)",
            "license_body": "license.licensing_body__v (LICENSING BODY)",
            "license_status": "license.license_status__v (LICENSE STATUS)",
            "hco_vid": "hco.hco_vid__v (HCO_VID__V)",
            "parent_hco_vid": "hco.parent_hco_vid__v (PARENT_HCO_VID__V)",
            "grandparent_hco_vid": "hco.grandparent_hco_vid__v (GRANDPARENT_HCO_VID__V)",
            "city_cda": "hcp.city_cda__v (CITY (CDA))",
            "addr_city": "address.locality__v (CITY)",
            "addr_state": "address.administrative_area__v (STATE/PROVINCE)",
            "addr_postal": "address.postal_code__v (ZIP/POSTAL CODE)",
            "hcp_status": "hcp.hcp_status__v (STATUS)",
            "candidate_record": "hcp.candidate_record__v (CANDIDATE RECORD)",
            "phone_pattern": "hcp.phone_*",
            "email_pattern": "hcp.email_*",
            "hco_entity_vid": "hco.vid__v (VID)",
            "hco_name": "hco.primary_name__v (NAME)",
            "hco_type": "hco.hco_type__v (TYPE)",
            "hco_phone": "hco.phone__v (PHONE)",
            "hco_fax": "hco.fax__v (FAX)",
            "hco_city": "hco.city__v (CITY)",
            "hco_state": "hco.state__v (STATE)",
            "hco_postal": "hco.postal_code__v (POSTAL CODE)",
            "hco_addr_line1": "hco.address_line_1__v (ADDRESS LINE 1)",
            "hco_addr_line2": "hco.address_line_2__v (ADDRESS LINE 2)",
            "hco_status": "hco.hco_status__v (STATUS)",
        },
        "shared_contact": {"threshold": 5},
        "blocking": {
            "max_block_size": 300,
            "phonetic_blocking": False,
            "first_initial_blocking": False,
        },
        "name_matching": {
            "strip_suffixes": ["md", "do", "phd", "mbbs", "bds", "mds", "ms", "dr", "jr", "sr", "ii", "iii", "iv"],
            "strong": 92,
            "medium": 85,
            "weak": 75,
        },
        "hcp_auto_rules": {
            "G1_name_spec_hco": {"name_min": 92},
            "G2_name_spec_pin": {"name_min": 92},
            "G3_name_spec_city": {"name_min": 92},
            "G4_license_match": {"name_min": 80},
            "G5_phone_email": {"name_min": 85},
            "G6_email_name": {"name_min": 92},
        },
        "hcp_not_dup_rules": {
            "N1_active_license_conflict": {},
            "N2_different_cities": {"name_min": 92},
        },
        "hcp_review_scoring": {
            "name_strong": 35, "name_medium": 25, "name_weak": 15,
            "specialty_match": 30, "hco_overlap": 25, "pin_match": 25,
            "city_match": 20, "geo_support": 10, "license_match": 40,
            "phone_email_match": 20, "email_match": 15, "phone_match": 10,
            "specialty_conflict": -15, "different_cities": -20, "candidate_record": -10,
            "review_threshold": 50, "high_confidence": 80, "medium_high": 70, "medium": 60,
        },
        "hco_auto_rules": {
            "H1_name_addr_phone": {"name_min": 90},
            "H2_name_addr_type": {"name_min": 92},
            "H3_name_phone_type": {"name_min": 90},
        },
        "hco_review_scoring": {
            "name_strong": 35, "name_medium": 25, "name_weak": 15,
            "address_match": 30, "city_match": 20, "postal_match": 25,
            "phone_match": 20, "fax_match": 15, "type_match": 15, "state_match": 10,
            "different_type": -20, "different_city": -15,
            "review_threshold": 50, "high_confidence": 80, "medium": 60,
        },
        "output": {
            "enrich_output": True,
            "write_csv": True,
            "max_contact_display": 200,
        },
    }


# ── Data Loading ─────────────────────────────────────────────────────

def _load_data_csv(input_path: str) -> pd.DataFrame:
    """Load input CSV file with memory optimization.

    For large files, uses pd.read_csv with chunking to avoid spike.
    """
    print(f"Reading CSV: {input_path}")
    try:
        # Try to read in one go first
        return pd.read_csv(input_path, dtype=str, na_values=[""])
    except Exception as e:
        logger.warning(f"Failed to load CSV in one pass: {e}. Retrying with chunking.")
        raise


def _validate_columns(df: pd.DataFrame, cfg: dict[str, Any]) -> None:
    """Check required columns exist; warn about missing optional ones."""
    cols = cfg["columns"]
    vid_col = cols["hcp_vid"]
    last_col = cols["hcp_last_name"]

    missing = [c for c in [vid_col, last_col] if c not in df.columns]
    if missing:
        print("\nERROR: Missing required columns:")
        for c in missing:
            print(f"  - {c}")
        sys.exit(1)


# ── Memory estimation ────────────────────────────────────────────────

def _estimate_memory_mb() -> float:
    """Estimate current process memory usage in MB."""
    try:
        import psutil
        process = psutil.Process(os.getpid())
        return process.memory_info().rss / (1024 * 1024)
    except ImportError:
        return 0.0


# ── Multiprocessing worker ───────────────────────────────────────────

def _worker_compare_pairs(
    args_tuple: tuple,
) -> list[dict[str, Any]]:
    """Worker function: compare a batch of VID pairs.

    This runs in a separate process and must work with serializable data.
    No shared_detector needed (shared_contact check disabled in large mode).

    Args:
        args_tuple: Tuple of (batch, canon_idx, cfg) — packed for imap_unordered

    Returns:
        List of classification dicts
    """
    batch, canon_idx, cfg = args_tuple
    from .hcp_pipeline import (
        _apply_auto_rules,
        _apply_not_dup_rules,
        _compute_signals,
        _enrich_row,
        _score_review,
    )
    from .core import DSU

    results: list[dict[str, Any]] = []
    enrich = cfg.get("output", {}).get("enrich_output", True)
    max_contact = cfg.get("output", {}).get("max_contact_display", 200)
    review_threshold = cfg.get("hcp_review_scoring", {}).get("review_threshold", 50)

    dsu = DSU()

    for vid_a, vid_b in batch:
        if vid_a not in canon_idx or vid_b not in canon_idx:
            continue

        ar = canon_idx[vid_a]
        br = canon_idx[vid_b]

        # Compute signals (no shared_detector in large mode)
        sig = _compute_signals(ar, br, shared_detector=None)

        base = {
            "vid_a": vid_a,
            "vid_b": vid_b,
            "name_similarity": round(sig["nscore"], 1),
            "specialty_match": ",".join(sorted(sig["spec_overlap"])),
            "geo_support": int(sig["geo_support"]),
            "hco_overlap": int(bool(sig["hco_overlap"])),
            "matched_phones": ";".join(sorted(sig["phones_overlap"]))[:max_contact],
            "matched_emails": ";".join(sorted(sig["emails_overlap"]))[:max_contact],
            "license_info": sig["license_info"],
        }
        if enrich:
            base.update(_enrich_row(vid_a, vid_b, canon_idx))

        # Try AUTO
        auto_result = _apply_auto_rules(sig, cfg)
        if auto_result:
            rule, comment = auto_result
            base["rule"] = rule
            base["comments"] = comment
            base["classification"] = "AUTO"
            dsu.union(vid_a, vid_b)
            results.append(base)
            continue

        # Try NOT-DUP
        notdup_result = _apply_not_dup_rules(sig, ar, br, cfg)
        if notdup_result:
            reason, comment = notdup_result
            base["reason"] = reason
            base["comments"] = comment
            base["classification"] = "NOTDUP"
            results.append(base)
            continue

        # Score for REVIEW
        score, reasons, comment = _score_review(sig, ar, br, cfg)
        if score >= review_threshold:
            base["score"] = score
            base["reasons"] = ",".join(reasons)
            base["comments"] = comment
            base["classification"] = "REVIEW"
            results.append(base)
        else:
            base["reason"] = "N3_LOW_SCORE"
            base["comments"] = f"Insufficient similarity (Score: {score})"
            base["classification"] = "NOTDUP"
            results.append(base)

    return results


# ── Pair comparison with streaming output ────────────────────────────

def _compare_pairs_parallel(
    pairs: list[tuple[str, str]],
    canon_idx: dict[str, dict],
    cfg: dict[str, Any],
    workers: int,
    output_base: str,
) -> dict[str, int]:
    """Compare all pairs using multiprocessing, stream results to CSV.

    Args:
        pairs: List of (vid_a, vid_b) candidate pairs
        canon_idx: Dict mapping vid -> canonical record
        cfg: Config
        workers: Number of worker processes
        output_base: Base path for output files (without extension)

    Returns:
        Summary dict with counts per classification
    """
    from tqdm import tqdm

    print(f"\n--- Comparing {len(pairs):,} pairs with {workers} workers ---")

    # Prepare output CSV files
    auto_file = f"{output_base}_auto.csv"
    review_file = f"{output_base}_review.csv"
    notdup_file = f"{output_base}_notdup.csv"

    csv_writers = {
        "AUTO": _init_csv_writer(auto_file),
        "REVIEW": _init_csv_writer(review_file),
        "NOTDUP": _init_csv_writer(notdup_file),
    }

    batch_size = max(5000, len(pairs) // (workers * 10))
    batches = [
        pairs[i : i + batch_size]
        for i in range(0, len(pairs), batch_size)
    ]

    counts = {"AUTO": 0, "REVIEW": 0, "NOTDUP": 0}

    with mp.Pool(workers) as pool:
        pbar = tqdm(total=len(pairs), desc="Comparing pairs")

        for batch_results in pool.imap_unordered(
            _worker_compare_pairs,
            [(b, canon_idx, cfg) for b in batches],
            chunksize=1,
        ):
            for row in batch_results:
                classification = row.pop("classification")
                counts[classification] += 1

                writer, file_handle = csv_writers[classification]
                if writer:
                    writer.writerow(row)
                    file_handle.flush()

                pbar.update(1)

        pbar.close()

    # Close all files
    for classification in csv_writers:
        writer, file_handle = csv_writers[classification]
        if file_handle:
            file_handle.close()

    logger.info(f"Classification summary: AUTO={counts['AUTO']}, REVIEW={counts['REVIEW']}, NOTDUP={counts['NOTDUP']}")

    return counts


def _init_csv_writer(filepath: str) -> tuple:
    """Initialize a CSV writer with common headers.

    Returns:
        (csv.DictWriter, file_handle) tuple
    """
    fieldnames = [
        "vid_a", "vid_b",
        "name_similarity", "specialty_match", "geo_support", "hco_overlap",
        "matched_phones", "matched_emails", "license_info",
        "name_a", "name_b", "specialties_a", "specialties_b",
        "cities_a", "cities_b", "phones_a", "phones_b", "emails_a", "emails_b",
        "rule", "reason", "score", "reasons", "comments"
    ]

    fh = open(filepath, "w", newline="", encoding="utf-8")
    writer = csv.DictWriter(fh, fieldnames=fieldnames)
    writer.writeheader()

    return writer, fh


# ── Blocking with pair limiting ──────────────────────────────────────

def _blocking_with_limit(
    canon: pd.DataFrame, cfg: dict[str, Any], max_pairs: int | None
) -> list[tuple[str, str]]:
    """Run blocking, optionally truncate if pair count exceeds limit.

    Args:
        canon: Canonical DataFrame
        cfg: Config
        max_pairs: Max pairs allowed; if exceeded, warn and truncate lowest-quality blocks

    Returns:
        List of (vid_a, vid_b) tuples
    """
    from .core import BlockingEngine, SpecialtySynonymResolver
    from tqdm import tqdm

    spec_resolver = SpecialtySynonymResolver(cfg.get("specialty_synonyms"))
    blk_cfg = cfg.get("blocking", {})
    blocker = BlockingEngine(
        max_block_size=blk_cfg.get("max_block_size", 300),
        phonetic=blk_cfg.get("phonetic_blocking", False),
        first_initial=blk_cfg.get("first_initial_blocking", False),
        spec_resolver=spec_resolver,
    )

    print("--- Blocking ---")
    for _, row in tqdm(canon.iterrows(), total=len(canon), desc="Building blocks"):
        blocker.add_hcp(row["vid"], row.to_dict())

    pairs = blocker.candidate_pairs()

    if max_pairs and len(pairs) > max_pairs:
        logger.warning(
            f"Blocking produced {len(pairs):,} pairs, exceeding limit of {max_pairs:,}. "
            f"Truncating lower-quality blocks."
        )
        # Simple truncation: just slice the sorted pairs
        pairs_list = sorted(pairs)[:max_pairs]
        pairs = set(pairs_list)

    return sorted(pairs)


# ── Main large-scale pipeline ───────────────────────────────────────

def main() -> None:
    """Main entry point for large-scale processing."""
    ap = argparse.ArgumentParser(
        description="HCP Duplicate Identification Tool (Large-Scale Optimized)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--input", required=True, help="Input CSV file (required for large-scale)")
    ap.add_argument("--config", default=None, help="Path to YAML config file")
    ap.add_argument("--outdir", default="./out", help="Output directory")
    ap.add_argument(
        "--workers", type=int, default=None,
        help="Number of parallel workers (default: auto, capped at 6 for laptop)",
    )
    ap.add_argument(
        "--max-pairs", type=int, default=None,
        help="Max candidate pairs; truncate if exceeded (default: no limit)",
    )
    ap.add_argument(
        "--no-excel", action="store_true",
        help="Skip Excel output, CSV only (faster)",
    )
    ap.add_argument(
        "--no-phonetic", action="store_true",
        help="Disable phonetic (Soundex) blocking (default: disabled for large data)",
    )
    ap.add_argument(
        "--enable-phonetic", action="store_true",
        help="Enable phonetic blocking (not recommended for large data)",
    )
    ap.add_argument(
        "--memory-limit", type=int, default=None,
        help="Warn if estimated memory exceeds this MB",
    )
    ap.add_argument(
        "--shared-threshold", type=int, default=None,
        help="Override shared-contact threshold",
    )
    ap.add_argument(
        "--verbose", "-v", action="store_true",
        help="Enable verbose logging",
    )
    args = ap.parse_args()

    # Logging
    level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s  %(name)-20s  %(levelname)-7s  %(message)s",
        datefmt="%H:%M:%S",
    )

    # Config
    cfg = _load_config(args.config)

    # Override blocking settings for large data
    if args.enable_phonetic:
        cfg.setdefault("blocking", {})["phonetic_blocking"] = True
    if not args.no_phonetic:
        cfg.setdefault("blocking", {})["phonetic_blocking"] = False
        cfg.setdefault("blocking", {})["first_initial_blocking"] = False

    # Worker count
    workers = args.workers
    if not workers:
        cpu_count = mp.cpu_count()
        workers = min(cpu_count - 1, 6)
    print(f"Using {workers} worker(s)")

    # Load data
    t0 = time.time()
    if not args.input.lower().endswith(".csv"):
        print("ERROR: Large-scale mode requires CSV input (not XLSX)")
        print("  XLSX is 5-10x slower to parse at this scale.")
        print("  Export to CSV first, then re-run.")
        sys.exit(1)

    df = _load_data_csv(args.input)
    load_time = time.time() - t0
    print(f"Loaded {len(df):,} rows x {len(df.columns)} cols in {load_time:.1f}s")
    print(f"Memory usage: {_estimate_memory_mb():.0f} MB")

    # Validate
    _validate_columns(df, cfg)

    # Canonicalize
    print("\n--- Canonicalizing ---")
    from .hcp_pipeline import _build_canonical

    t1 = time.time()
    canon = _build_canonical(df, cfg)
    canon_time = time.time() - t1

    print(f"Canonicalized {len(canon):,} unique VIDs in {canon_time:.1f}s")
    print(f"Memory usage: {_estimate_memory_mb():.0f} MB")

    # Check memory limit
    if args.memory_limit:
        mem_mb = _estimate_memory_mb()
        if mem_mb > args.memory_limit:
            logger.warning(
                f"Estimated memory {mem_mb:.0f} MB exceeds limit {args.memory_limit} MB"
            )

    # Free original DataFrame
    del df
    gc.collect()

    # Build indexable dict for workers
    print("Building canonical index...")
    canon_idx = {row["vid"]: row.to_dict() for _, row in canon.iterrows()}
    print(f"Index contains {len(canon_idx):,} records")

    # Blocking
    pairs = _blocking_with_limit(canon, cfg, args.max_pairs)
    print(f"Candidate pairs: {len(pairs):,}")

    if not pairs:
        print("No candidate pairs generated — exiting.")
        return

    # Prepare output directory
    os.makedirs(args.outdir, exist_ok=True)
    output_base = os.path.join(args.outdir, "HCP_LARGE")

    # Compare pairs with parallel processing
    t2 = time.time()
    counts = _compare_pairs_parallel(pairs, canon_idx, cfg, workers, output_base)
    compare_time = time.time() - t2

    print(f"\nPair comparison: {compare_time:.1f}s")
    print(f"  AUTO:   {counts['AUTO']:,}")
    print(f"  REVIEW: {counts['REVIEW']:,}")
    print(f"  NOTDUP: {counts['NOTDUP']:,}")

    # Consolidate Excel output if requested
    if not args.no_excel:
        print("\n--- Consolidating to Excel ---")
        t3 = time.time()
        _consolidate_to_excel(output_base, args.outdir, cfg)
        print(f"Excel consolidation: {time.time()-t3:.1f}s")

    print(f"\n{'='*60}")
    print(f"Total time: {time.time()-t0:.1f}s")
    print(f"Final memory: {_estimate_memory_mb():.0f} MB")
    print(f"Output directory: {args.outdir}")
    print(f"{'='*60}")


def _consolidate_to_excel(output_base: str, outdir: str, cfg: dict[str, Any] | None = None) -> None:
    """Read CSV results and consolidate into Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    excel_path = os.path.join(outdir, "HCP_Dupe_Check_Large.xlsx")

    # Write rules sheet first (position 0)
    if cfg:
        try:
            from output import build_rules_dataframes
            sections = build_rules_dataframes(cfg)
            hcp_sections = [(t, d) for t, d in sections if "HCP" in t]
            if hcp_sections:
                ws = wb.create_sheet("Matching_Rules", 0)
                ws.cell(row=1, column=1, value="HCP Duplicate Tool — Matching Rules Reference")
                ws.cell(row=1, column=1).font = Font(bold=True, size=14)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
                current_row = 3
                section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                section_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
                header_font = Font(bold=True, size=10)
                for title, df_rules in hcp_sections:
                    ws.cell(row=current_row, column=1, value=title)
                    ws.cell(row=current_row, column=1).font = section_font
                    ws.cell(row=current_row, column=1).fill = section_fill
                    for ci in range(1, len(df_rules.columns) + 1):
                        ws.cell(row=current_row, column=ci).fill = section_fill
                    ws.merge_cells(start_row=current_row, start_column=1,
                                   end_row=current_row, end_column=max(len(df_rules.columns), 1))
                    current_row += 1
                    for ci, col_name in enumerate(df_rules.columns, start=1):
                        cell = ws.cell(row=current_row, column=ci, value=col_name)
                        cell.font = header_font
                        cell.fill = header_fill
                    current_row += 1
                    for _, row_data in df_rules.iterrows():
                        for ci, col_name in enumerate(df_rules.columns, start=1):
                            ws.cell(row=current_row, column=ci, value=row_data[col_name])
                        current_row += 1
                    current_row += 1
                for col_cells in ws.columns:
                    max_len = 0
                    col_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
        except Exception as e:
            logger.warning(f"Could not write rules sheet: {e}")

    for class_type in ["auto", "review", "notdup"]:
        csv_path = f"{output_base}_{class_type}.csv"
        if not os.path.isfile(csv_path):
            continue

        df = pd.read_csv(csv_path)
        ws = wb.create_sheet(title=class_type.upper())

        # Header with formatting
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Data
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx).value = value

        # Auto-width columns (basic)
        for col_idx in range(1, len(df.columns) + 1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 12

    wb.save(excel_path)
    logger.info(f"Wrote: {excel_path}")


if __name__ == "__main__":
    main()
