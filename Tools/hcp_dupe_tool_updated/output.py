"""
output.py — Write results to Excel (multi-sheet) and CSV files.

Produces reviewer-friendly output with enrichment columns so
that a human reviewer can act directly from the sheet without
cross-referencing VIDs.
"""

from __future__ import annotations

import logging
import os
from typing import Any

import pandas as pd
from openpyxl.utils import get_column_letter

logger = logging.getLogger("dupe_tool.output")


# ── VID Column Formatting ─────────────────────────────────────────────

# Column names (case-insensitive) that contain VID values — 18-digit IDs
# that Excel mangles into scientific notation if stored as numbers.
_VID_COLUMN_KEYWORDS = {"vid", "vid_a", "vid_b", "paired_with", "matched_vid",
                        "hco_vid", "parent_hco_vid", "grandparent_hco_vid"}


def _format_vid_columns_as_text(writer: pd.ExcelWriter, sheet_name: str,
                                 df: pd.DataFrame) -> None:
    """Format VID columns as text in an already-written Excel sheet.

    Must be called *after* ``df.to_excel(writer, sheet_name=…)`` so that
    the worksheet exists.  Rewrites every cell in VID columns as a string
    value so Excel never coerces them to floats / scientific notation.
    """
    if df.empty:
        return
    ws = writer.sheets[sheet_name]
    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name.lower().strip() in _VID_COLUMN_KEYWORDS:
            col_letter = get_column_letter(col_idx)
            for row_idx in range(2, len(df) + 2):          # row 1 = header
                cell = ws[f"{col_letter}{row_idx}"]
                if cell.value is not None:
                    cell.value = str(cell.value)
                    cell.number_format = "@"                # text format


# ── Rules Sheet Generator ─────────────────────────────────────────────

def build_rules_dataframes(cfg: dict[str, Any]) -> list[tuple[str, pd.DataFrame]]:
    """Build DataFrames describing all matching rules from the config.

    Returns a list of (section_title, DataFrame) tuples, ready to be
    written row-by-row into a single "Matching_Rules" Excel sheet.
    """
    # Built-in rule metadata (always available even without YAML)
    _HCP_AUTO_META = {
        "G1_name_spec_hco": {
            "requires": "specialty_overlap, hco_overlap",
            "description": "Name + Specialty + HCO affiliation match",
        },
        "G2_name_spec_pin": {
            "requires": "specialty_overlap, pin_overlap",
            "description": "Name + Specialty + Postal code match",
        },
        "G3_name_spec_city": {
            "requires": "specialty_overlap, city_overlap",
            "description": "Name + Specialty + City match",
        },
        "G4_license_match": {
            "requires": "active_license_match",
            "description": "Active license match + reasonable name",
        },
        "G5_phone_email": {
            "requires": "phone_overlap, email_overlap, not_shared_contact",
            "description": "Phone + Email match (non-shared contacts)",
        },
        "G6_email_name": {
            "requires": "email_overlap, not_shared_contact",
            "description": "Email + Strong name match",
        },
    }
    _HCP_NOTDUP_META = {
        "N1_active_license_conflict": {
            "requires": "active_license_conflict",
            "description": "Both have active licenses but none overlap",
            "exception_if": "phone_and_email_overlap",
        },
        "N2_different_cities": {
            "requires": "specialty_overlap, cities_explicitly_different",
            "description": "Same name + specialty but explicitly different cities",
            "exception_if": "phone_and_email_overlap, license_match",
        },
    }
    _HCO_AUTO_META = {
        "H1_name_addr_phone": {
            "requires": "address_overlap, phone_or_fax_overlap",
            "description": "Name + Address + Phone/Fax match",
        },
        "H2_name_addr_type": {
            "requires": "address_overlap, type_match",
            "description": "Name + Address + Same type",
        },
        "H3_name_phone_type": {
            "requires": "phone_or_fax_overlap, type_match",
            "description": "Name + Phone/Fax + Same type",
        },
    }

    frames: list[tuple[str, pd.DataFrame]] = []

    # ── HCP Auto-Merge Rules ──────────────────────────────────
    hcp_auto = cfg.get("hcp_auto_rules", {})
    rows = []
    for rule_id, rule in hcp_auto.items():
        meta = _HCP_AUTO_META.get(rule_id, {})
        name_min = rule.get("name_min", "")
        requires = ", ".join(rule.get("requires", [])) or meta.get("requires", "")
        desc = rule.get("description", "") or meta.get("description", "")
        note = ""
        if rule_id == "G3_name_spec_city":
            note = (
                "Sub-rules: G3a (+ phone → auto), G3b (+ license → auto), "
                "G3c (uncommon name → auto), G3d (common name → REVIEW)"
            )
        rows.append({
            "Rule": rule_id,
            "Name Similarity Min": name_min,
            "Required Signals": requires,
            "Description": desc,
            "Verdict": "AUTO-MERGE (Confirmed Duplicate)",
            "Notes": note,
        })
    if rows:
        frames.append(("HCP AUTO-MERGE RULES", pd.DataFrame(rows)))

    # ── HCP Not-Duplicate Rules ───────────────────────────────
    hcp_notdup = cfg.get("hcp_not_dup_rules", {})
    rows = []
    for rule_id, rule in hcp_notdup.items():
        meta = _HCP_NOTDUP_META.get(rule_id, {})
        requires = ", ".join(rule.get("requires", [])) or meta.get("requires", "")
        exc = ", ".join(rule.get("exception_if", [])) or meta.get("exception_if", "")
        desc = rule.get("description", "") or meta.get("description", "")
        name_min = rule.get("name_min", "")
        rows.append({
            "Rule": rule_id,
            "Name Similarity Min": name_min,
            "Required Signals": requires,
            "Description": desc,
            "Verdict": "NOT-DUPLICATE",
            "Exceptions (→ REVIEW instead)": exc,
        })
    if rows:
        frames.append(("HCP NOT-DUPLICATE RULES", pd.DataFrame(rows)))

    # ── HCP Review Scoring Weights ────────────────────────────
    hcp_scoring = cfg.get("hcp_review_scoring", {})
    pos_rows = []
    neg_rows = []
    meta_rows = []
    scoring_meta_keys = {"review_threshold", "high_confidence", "medium_high", "medium"}
    for key, val in hcp_scoring.items():
        label = key.replace("_", " ").title()
        if key in scoring_meta_keys:
            meta_rows.append({"Parameter": label, "Value": val})
        elif isinstance(val, (int, float)) and val < 0:
            neg_rows.append({"Signal": label, "Weight": val})
        else:
            pos_rows.append({"Signal": label, "Weight": val})
    scoring_rows = []
    for r in pos_rows:
        scoring_rows.append({"Signal": r["Signal"], "Weight": r["Weight"], "Type": "Positive"})
    for r in neg_rows:
        scoring_rows.append({"Signal": r["Signal"], "Weight": r["Weight"], "Type": "Negative (penalty)"})
    if scoring_rows:
        frames.append(("HCP REVIEW SCORING WEIGHTS", pd.DataFrame(scoring_rows)))
    if meta_rows:
        frames.append(("HCP REVIEW THRESHOLDS", pd.DataFrame(meta_rows)))

    # ── HCO Auto-Merge Rules ─────────────────────────────────
    hco_auto = cfg.get("hco_auto_rules", {})
    rows = []
    for rule_id, rule in hco_auto.items():
        meta = _HCO_AUTO_META.get(rule_id, {})
        name_min = rule.get("name_min", "")
        requires = ", ".join(rule.get("requires", [])) or meta.get("requires", "")
        desc = rule.get("description", "") or meta.get("description", "")
        rows.append({
            "Rule": rule_id,
            "Name Similarity Min": name_min,
            "Required Signals": requires,
            "Description": desc,
            "Verdict": "AUTO-MERGE (Confirmed Duplicate)",
        })
    if rows:
        frames.append(("HCO AUTO-MERGE RULES", pd.DataFrame(rows)))

    # ── HCO Review Scoring Weights ────────────────────────────
    hco_scoring = cfg.get("hco_review_scoring", {})
    hco_meta_keys = {"review_threshold", "high_confidence", "medium"}
    scoring_rows = []
    meta_rows = []
    for key, val in hco_scoring.items():
        label = key.replace("_", " ").title()
        if key in hco_meta_keys:
            meta_rows.append({"Parameter": label, "Value": val})
        elif isinstance(val, (int, float)) and val < 0:
            scoring_rows.append({"Signal": label, "Weight": val, "Type": "Negative (penalty)"})
        else:
            scoring_rows.append({"Signal": label, "Weight": val, "Type": "Positive"})
    if scoring_rows:
        frames.append(("HCO REVIEW SCORING WEIGHTS", pd.DataFrame(scoring_rows)))
    if meta_rows:
        frames.append(("HCO REVIEW THRESHOLDS", pd.DataFrame(meta_rows)))

    return frames


def write_rules_sheet(
    writer: pd.ExcelWriter,
    cfg: dict[str, Any],
    entity_filter: str | None = None,
) -> None:
    """Write a 'Matching_Rules' sheet as the FIRST sheet in an ExcelWriter.

    Args:
        writer: An open pd.ExcelWriter (openpyxl engine).
        cfg: The tool configuration dict.
        entity_filter: If "hcp", only write HCP rules. If "hco", only HCO.
                       If None, write all rules.
    """
    sections = build_rules_dataframes(cfg)

    if entity_filter:
        tag = entity_filter.upper()
        sections = [(title, df) for title, df in sections if tag in title]

    if not sections:
        return

    # We build a single combined DataFrame with section headers as separator rows
    ws_name = "Matching_Rules"
    combined_rows: list[dict] = []
    for title, df in sections:
        # Section header row
        combined_rows.append({"Section": f"── {title} ──"})
        # Column header row
        header_row = {col: col for col in df.columns}
        header_row["Section"] = ""
        combined_rows.append(header_row)
        # Data rows
        for _, row in df.iterrows():
            data = row.to_dict()
            data["Section"] = ""
            combined_rows.append(data)
        # Blank separator
        combined_rows.append({})

    combined_df = pd.DataFrame(combined_rows)

    # Use openpyxl directly for better formatting
    try:
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = writer.book
        ws = wb.create_sheet(ws_name, 0)  # Position 0 = first sheet

        # Title row
        ws.cell(row=1, column=1, value="HCP / HCO Duplicate Tool — Matching Rules Reference")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        current_row = 3
        section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        header_font = Font(bold=True, size=10)

        for title, df in sections:
            # Section header
            ws.cell(row=current_row, column=1, value=title)
            ws.cell(row=current_row, column=1).font = section_font
            ws.cell(row=current_row, column=1).fill = section_fill
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=current_row, column=col_idx).fill = section_fill
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=max(len(df.columns), 1),
            )
            current_row += 1

            # Column headers
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
            current_row += 1

            # Data rows
            for _, row in df.iterrows():
                for col_idx, col_name in enumerate(df.columns, start=1):
                    val = row[col_name]
                    ws.cell(row=current_row, column=col_idx, value=val)
                current_row += 1

            current_row += 1  # blank separator

        # Auto-width columns (iterate by column index to avoid merged cell issues)
        max_col = ws.max_column or 1
        for col_idx in range(1, max_col + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and not isinstance(cell, type(None)):
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    except Exception as e:
        logger.warning("Could not format rules sheet: %s", e)


def write_results(
    hcp_results: dict[str, pd.DataFrame],
    hco_results: dict[str, pd.DataFrame],
    outdir: str,
    cfg: dict[str, Any],
) -> list[str]:
    """Write all result DataFrames to Excel and (optionally) CSV.

    Returns a list of file paths written.
    """

    os.makedirs(outdir, exist_ok=True)
    write_csv = cfg.get("output", {}).get("write_csv", True)
    written: list[str] = []

    # ── HCP Outputs ──────────────────────────────────────────────────
    hcp_main = os.path.join(outdir, "HCP_Dupe_Check.xlsx")
    hcp_canon_path = os.path.join(outdir, "HCP_1row_per_VID.xlsx")

    hcp_canonical = hcp_results.get("hcp_canonical", pd.DataFrame())
    hcp_auto = hcp_results.get("hcp_auto", pd.DataFrame())
    hcp_review = hcp_results.get("hcp_review", pd.DataFrame())
    hcp_notdup = hcp_results.get("hcp_notdup", pd.DataFrame())
    hcp_unique = hcp_results.get("hcp_unique", pd.DataFrame())
    hcp_clusters = hcp_results.get("hcp_clusters", pd.DataFrame())
    hcp_shared = hcp_results.get("hcp_shared", pd.DataFrame())
    hcp_summary = hcp_results.get("hcp_summary", pd.DataFrame())

    # Canonical standalone
    if not hcp_canonical.empty:
        with pd.ExcelWriter(hcp_canon_path, engine="openpyxl") as w:
            hcp_canonical.to_excel(w, index=False, sheet_name="HCP_Canonical")
            _format_vid_columns_as_text(w, "HCP_Canonical", hcp_canonical)
        written.append(hcp_canon_path)
        logger.info("Wrote: %s", hcp_canon_path)

    # Main workbook
    with pd.ExcelWriter(hcp_main, engine="openpyxl") as w:
        write_rules_sheet(w, cfg, entity_filter="hcp")
        if not hcp_summary.empty:
            hcp_summary.to_excel(w, index=False, sheet_name="Summary")
        if not hcp_auto.empty:
            hcp_auto.to_excel(w, index=False, sheet_name="AUTO_MERGE")
            _format_vid_columns_as_text(w, "AUTO_MERGE", hcp_auto)
        else:
            pd.DataFrame({"info": ["No auto-merge pairs found"]}).to_excel(
                w, index=False, sheet_name="AUTO_MERGE"
            )
        if not hcp_review.empty:
            hcp_review.to_excel(w, index=False, sheet_name="REVIEW")
            _format_vid_columns_as_text(w, "REVIEW", hcp_review)
        else:
            pd.DataFrame({"info": ["No review pairs found"]}).to_excel(
                w, index=False, sheet_name="REVIEW"
            )
        if not hcp_notdup.empty:
            hcp_notdup.to_excel(w, index=False, sheet_name="NOT_DUP")
            _format_vid_columns_as_text(w, "NOT_DUP", hcp_notdup)
        if not hcp_clusters.empty:
            hcp_clusters.to_excel(w, index=False, sheet_name="AUTO_CLUSTERS")
            _format_vid_columns_as_text(w, "AUTO_CLUSTERS", hcp_clusters)
        if not hcp_shared.empty:
            hcp_shared.to_excel(w, index=False, sheet_name="Shared_Contacts")
            _format_vid_columns_as_text(w, "Shared_Contacts", hcp_shared)
        if not hcp_unique.empty:
            hcp_unique.to_excel(w, index=False, sheet_name="UNIQUE")
            _format_vid_columns_as_text(w, "UNIQUE", hcp_unique)
    written.append(hcp_main)
    logger.info("Wrote: %s", hcp_main)

    # CSV
    if write_csv:
        csv_map = {
            "HCP_AUTO_MERGE.csv": hcp_auto,
            "HCP_REVIEW.csv": hcp_review,
            "HCP_NOT_DUP.csv": hcp_notdup,
            "HCP_UNIQUE.csv": hcp_unique,
            "HCP_AUTO_CLUSTERS.csv": hcp_clusters,
            "HCP_Shared_Contacts.csv": hcp_shared,
        }
        for fname, frame in csv_map.items():
            if not frame.empty:
                path = os.path.join(outdir, fname)
                frame.to_csv(path, index=False)
                written.append(path)

    # ── HCO Outputs ──────────────────────────────────────────────────
    hco_canonical = hco_results.get("hco_canonical", pd.DataFrame())
    hco_auto = hco_results.get("hco_auto", pd.DataFrame())
    hco_review = hco_results.get("hco_review", pd.DataFrame())
    hco_notdup = hco_results.get("hco_notdup", pd.DataFrame())
    hco_unique = hco_results.get("hco_unique", pd.DataFrame())
    hco_clusters = hco_results.get("hco_clusters", pd.DataFrame())
    hco_summary = hco_results.get("hco_summary", pd.DataFrame())

    if not hco_canonical.empty:
        hco_main = os.path.join(outdir, "HCO_Dupe_Check.xlsx")
        hco_canon_path = os.path.join(outdir, "HCO_1row_per_VID.xlsx")

        with pd.ExcelWriter(hco_canon_path, engine="openpyxl") as w:
            hco_canonical.to_excel(w, index=False, sheet_name="HCO_Canonical")
            _format_vid_columns_as_text(w, "HCO_Canonical", hco_canonical)
        written.append(hco_canon_path)

        with pd.ExcelWriter(hco_main, engine="openpyxl") as w:
            write_rules_sheet(w, cfg, entity_filter="hco")
            if not hco_summary.empty:
                hco_summary.to_excel(w, index=False, sheet_name="Summary")
            if not hco_auto.empty:
                hco_auto.to_excel(w, index=False, sheet_name="AUTO_MERGE")
                _format_vid_columns_as_text(w, "AUTO_MERGE", hco_auto)
            else:
                pd.DataFrame({"info": ["No auto-merge pairs found"]}).to_excel(
                    w, index=False, sheet_name="AUTO_MERGE"
                )
            if not hco_review.empty:
                hco_review.to_excel(w, index=False, sheet_name="REVIEW")
                _format_vid_columns_as_text(w, "REVIEW", hco_review)
            else:
                pd.DataFrame({"info": ["No review pairs found"]}).to_excel(
                    w, index=False, sheet_name="REVIEW"
                )
            if not hco_notdup.empty:
                hco_notdup.to_excel(w, index=False, sheet_name="NOT_DUP")
                _format_vid_columns_as_text(w, "NOT_DUP", hco_notdup)
            if not hco_clusters.empty:
                hco_clusters.to_excel(w, index=False, sheet_name="AUTO_CLUSTERS")
                _format_vid_columns_as_text(w, "AUTO_CLUSTERS", hco_clusters)
            if not hco_unique.empty:
                hco_unique.to_excel(w, index=False, sheet_name="UNIQUE")
                _format_vid_columns_as_text(w, "UNIQUE", hco_unique)
        written.append(hco_main)
        logger.info("Wrote: %s", hco_main)

        if write_csv:
            csv_map = {
                "HCO_AUTO_MERGE.csv": hco_auto,
                "HCO_REVIEW.csv": hco_review,
                "HCO_NOT_DUP.csv": hco_notdup,
                "HCO_UNIQUE.csv": hco_unique,
                "HCO_AUTO_CLUSTERS.csv": hco_clusters,
            }
            for fname, frame in csv_map.items():
                if not frame.empty:
                    path = os.path.join(outdir, fname)
                    frame.to_csv(path, index=False)
                    written.append(path)
    else:
        logger.info("No HCO records — skipping HCO output files.")

    return written


# ── Tagged Source Builder ─────────────────────────────────────────────

def build_tagged_source(
    results: dict[str, pd.DataFrame],
    entity_type: str = "hco",
) -> pd.DataFrame:
    """Build a simplified per-VID actionable view of dupe results.

    Returns a flat table with one row per VID-match combination:
      - vid:          source VID
      - name:         entity name (HCO name or HCP name)
      - status:       AUTO_MERGE | REVIEW | NOT_DUP
      - matched_vid:  the VID it was paired against
      - matched_name: the paired entity's name

    VIDs with multiple matches appear on multiple rows (easy to filter).
    Clean records (NO_MATCH) are excluded — only actionable items shown.
    Sorted by status priority (AUTO_MERGE first) then by VID.
    """
    prefix = entity_type  # "hco" or "hcp"

    canonical = results.get(f"{prefix}_canonical", pd.DataFrame())
    auto_df = results.get(f"{prefix}_auto", pd.DataFrame())
    review_df = results.get(f"{prefix}_review", pd.DataFrame())
    notdup_df = results.get(f"{prefix}_notdup", pd.DataFrame())
    unique_df = results.get(f"{prefix}_unique", pd.DataFrame())

    if canonical.empty:
        return pd.DataFrame()

    # Build vid → display name lookup from canonical
    name_col = "display_name" if "display_name" in canonical.columns else "name_norm"
    vid_name: dict[str, str] = {}
    if name_col in canonical.columns:
        for _, row in canonical[["vid", name_col]].iterrows():
            vid_name[str(row["vid"])] = str(row[name_col])

    # Collect rows from each pair sheet
    rows: list[dict[str, str]] = []
    status_map = [
        (auto_df,   "AUTO_MERGE"),
        (review_df, "REVIEW"),
        (notdup_df, "NOT_DUP"),
    ]

    seen_pairs: set[tuple[str, str, str]] = set()  # (vid, matched_vid, status)

    for pair_df, status in status_map:
        if pair_df.empty or "vid_a" not in pair_df.columns:
            continue
        # Use name columns from pair sheet if available, else fall back to canonical lookup
        has_name_a = "name_a" in pair_df.columns
        has_name_b = "name_b" in pair_df.columns

        for _, row in pair_df.iterrows():
            va, vb = str(row["vid_a"]), str(row["vid_b"])
            na = str(row["name_a"]) if has_name_a else vid_name.get(va, "")
            nb = str(row["name_b"]) if has_name_b else vid_name.get(vb, "")

            # Emit one row per side of the pair (A→B and B→A)
            for vid, name, m_vid, m_name in [(va, na, vb, nb), (vb, nb, va, na)]:
                key = (vid, m_vid, status)
                if key not in seen_pairs:
                    seen_pairs.add(key)
                    rows.append({
                        "vid": vid,
                        "name": name,
                        "status": status,
                        "matched_vid": m_vid,
                        "matched_name": m_name,
                    })

    # Add UNIQUE VIDs (no candidate pairs — confirmed unique)
    if not unique_df.empty and "vid" in unique_df.columns:
        for _, row in unique_df.iterrows():
            vid = str(row["vid"])
            name = str(row.get("name", vid_name.get(vid, "")))
            rows.append({
                "vid": vid,
                "name": name,
                "status": "UNIQUE",
                "matched_vid": "",
                "matched_name": "",
            })

    if not rows:
        return pd.DataFrame(columns=["vid", "name", "status", "matched_vid", "matched_name"])

    tagged = pd.DataFrame(rows)

    # Sort: AUTO_MERGE first, then REVIEW, then NOT_DUP, then UNIQUE — then by VID
    status_order = {"AUTO_MERGE": 0, "REVIEW": 1, "NOT_DUP": 2, "UNIQUE": 3}
    tagged["_sort"] = tagged["status"].map(status_order)
    tagged = tagged.sort_values(["_sort", "vid", "matched_vid"]).drop(columns=["_sort"])
    tagged = tagged.reset_index(drop=True)

    return tagged


def write_pdr_results(
    pdr_df: pd.DataFrame,
    ref_vid_count: int,
    elapsed_seconds: float,
    outdir: str,
    cfg: dict[str, Any] | None = None,
) -> list[str]:
    """Write PDR pre-screening results to Excel and CSV.

    Returns a list of file paths written.
    """
    if pdr_df.empty:
        logger.info("No PDR results to write.")
        return []

    os.makedirs(outdir, exist_ok=True)
    written: list[str] = []

    dup_ct = int((pdr_df["pdr_verdict"] == "LIKELY_DUP").sum())
    poss_ct = int((pdr_df["pdr_verdict"] == "POSSIBLE_MATCH").sum())
    clean_ct = int((pdr_df["pdr_verdict"] == "CLEAN").sum())

    # Excel workbook
    xlsx_path = os.path.join(outdir, "PDR_PreScreen.xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        if cfg:
            write_rules_sheet(w, cfg, entity_filter="hcp")
        summary = pd.DataFrame([{
            "total_pdr_records": len(pdr_df),
            "likely_duplicates": dup_ct,
            "possible_matches": poss_ct,
            "clean_records": clean_ct,
            "reference_db_vids": ref_vid_count,
            "screening_time_seconds": round(elapsed_seconds, 1),
        }])
        summary.to_excel(w, index=False, sheet_name="Summary")
        pdr_df.to_excel(w, index=False, sheet_name="All_Results")
        if dup_ct > 0:
            pdr_df[pdr_df["pdr_verdict"] == "LIKELY_DUP"].to_excel(
                w, index=False, sheet_name="Likely_Duplicates"
            )
        if poss_ct > 0:
            pdr_df[pdr_df["pdr_verdict"] == "POSSIBLE_MATCH"].to_excel(
                w, index=False, sheet_name="Possible_Matches"
            )
        if clean_ct > 0:
            pdr_df[pdr_df["pdr_verdict"] == "CLEAN"].to_excel(
                w, index=False, sheet_name="Clean"
            )
    written.append(xlsx_path)
    logger.info("Wrote: %s", xlsx_path)

    # CSV — all results
    csv_path = os.path.join(outdir, "PDR_PreScreen_All.csv")
    pdr_df.to_csv(csv_path, index=False)
    written.append(csv_path)

    # CSV — flagged only
    flagged = pdr_df[pdr_df["pdr_verdict"] != "CLEAN"]
    if not flagged.empty:
        flagged_path = os.path.join(outdir, "PDR_PreScreen_Flagged.csv")
        flagged.to_csv(flagged_path, index=False)
        written.append(flagged_path)

    return written
