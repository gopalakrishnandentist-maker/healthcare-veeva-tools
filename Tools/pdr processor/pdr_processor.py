"""
AI-Augmented PDR Processing Tool
=================================
Processes Proactive Data Change Requests (PDRs) for HCP data validation.
Runs 100% locally — no cloud, no API calls, no data leaves this machine.

Usage:
    python pdr_processor.py --pdr <pdr_file.xlsx> --nwk <nwk_file.xlsx> --output <output.xlsx>

Or import and use programmatically:
    from pdr_processor import PDRProcessor
    processor = PDRProcessor(pdr_path, nwk_path)
    results = processor.run()
    processor.export(results, output_path)
"""

import pandas as pd
import re
import argparse
import os
import sys
from difflib import SequenceMatcher
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIGURATION
# =============================================================================

class Config:
    """All tunable parameters in one place."""

    # Fuzzy matching thresholds (0-100)
    EXACT_MATCH_THRESHOLD = 92       # Above this = auto-reject R-10075
    PARTIAL_MATCH_THRESHOLD = 70     # Between this and EXACT = flag for review
    NO_MATCH_THRESHOLD = 70          # Below this = no match found

    # Scope rules — out-of-scope specialties/degrees (case-insensitive substrings)
    OUT_OF_SCOPE_SPECIALTIES = [
        "dentist", "dental", "bds", "mds", "orthodont",
        "pharmacy", "pharmacist", "pharm.d", "b.pharm", "m.pharm", "d.pharm",
        "homeopath", "bhms", "homoeopath",
        "ayurved", "bams", "ayush",
        "unani", "bums",
        "siddha", "bsms",
        "naturopath", "bnys",
        "veterinar", "bvsc", "b.v.sc",
        "physiotherap", "bpt", "mpt",
        "nursing", "b.sc nursing", "gnm",
        "optometr",
        "yoga",
    ]

    # In-scope degree indicators
    IN_SCOPE_DEGREES = [
        "mbbs", "md", "ms", "dm", "mch", "dnb", "diploma",
        "frcs", "mrcp", "frcp", "frcog",
    ]

    # Business professional titles (in-scope even without MBBS)
    BUSINESS_TITLES = [
        "ceo", "coo", "cfo", "cmo", "cto",
        "director", "manager", "administrator",
        "president", "vice president", "vp",
        "chairman", "chairperson",
        "head of", "chief",
    ]

    # Junk data patterns
    JUNK_PATTERNS = [
        r"(?i)review[s]?\s+(by|from|on)",       # Scraped reviews
        r"(?i)author[s]?\s*:",                    # Author listings
        r"(?i)(rating|stars?|feedback)",           # Rating text
        r"(?i)(association|society|council|forum|federation)\b",  # Non-individual entities
        r"\d{10,}",                               # Long number strings (not phone)
    ]

    # Name validation
    MIN_NAME_LENGTH = 2
    MAX_NAME_LENGTH = 60


# =============================================================================
# STEWARD NOTE PARSERS
# =============================================================================

class NoteParser:
    """Parses structured steward notes from PDR data."""

    @staticmethod
    def detect_format(note):
        """Detect whether note is vod.manual or vod.omnia format."""
        if not note or not isinstance(note, str):
            return "unknown"
        if "// City:" in note or "// HCO_VID:" in note or "// HCP_Name:" in note:
            return "manual"
        if "** parenthco vid:" in note.lower() or "source url:" in note.lower():
            return "omnia"
        return "unknown"

    @staticmethod
    def parse_manual(note):
        """Parse vod.manual format: // Key: Value // Key: Value ..."""
        result = {
            "city": None,
            "hco_vid": None,
            "hco_name": None,
            "hco_url": None,
            "hcp_name": None,
            "specialty": None,
            "qualifications": None,
            "scope_status": None,
            "source_for_listing": None,
            "department": None,
            "parse_format": "manual",
        }
        if not note or not isinstance(note, str):
            return result

        # Clean up the note
        note_clean = note.strip().strip('"')

        # Extract key-value pairs using // delimiter
        patterns = {
            "city": r"//\s*City:\s*([^/]+?)(?://|Created|$)",
            "hco_vid": r"//\s*HCO_VID:\s*(\d+)",
            "hco_name": r"//\s*HCO_Name:\s*([^/]+?)(?://|Created|$)",
            "hco_url": r"//\s*HCO_URL:\s*(\S+?)(?=\s+//|\s+Created|\s*$)",
            "hcp_name": r"//\s*HCP_Name:\s*([^/]+?)(?://|Created|$)",
            "scope_status": r"//\s*Scope_status:\s*([^/]+?)(?://|Created|$)",
            "source_for_listing": r"//\s*source_for_hcp_listing:\s*([^/]+?)(?://|Created|$)",
        }

        for key, pattern in patterns.items():
            match = re.search(pattern, note_clean, re.IGNORECASE)
            if match:
                result[key] = match.group(1).strip()

        return result

    @staticmethod
    def parse_omnia(note):
        """Parse vod.omnia format: key: value ** key: value ..."""
        result = {
            "city": None,
            "hco_vid": None,
            "hco_name": None,
            "hco_url": None,
            "hcp_name": None,
            "specialty": None,
            "qualifications": None,
            "scope_status": None,
            "source_for_listing": None,
            "department": None,
            "first_name": None,
            "last_name": None,
            "middle_name": None,
            "salutation": None,
            "title": None,
            "parse_format": "omnia",
        }
        if not note or not isinstance(note, str):
            return result

        note_clean = note.strip().strip('"')

        # Source URL (appears at start)
        url_match = re.search(r"source url:\s*(\S+?)(?:\s*\*\*|$)", note_clean, re.IGNORECASE)
        if url_match:
            result["hco_url"] = url_match.group(1).strip()

        # Key-value pairs with ** delimiter
        patterns = {
            "hco_vid": r"\*\*\s*parenthco vid:\s*(\d+)",
            "hcp_name": r"\*\*\s*Name:\s*([^*]+?)(?:\*\*|Created|$)",
            "first_name": r"\*\*\s*First Name:\s*([^*]+?)(?:\*\*|Created|$)",
            "last_name": r"\*\*\s*Last Name:\s*([^*]+?)(?:\*\*|Created|$)",
            "middle_name": r"\*\*\s*Middle Names?:\s*([^*]+?)(?:\*\*|Created|$)",
            "salutation": r"\*\*\s*Salutation:\s*([^*]+?)(?:\*\*|Created|$)",
            "title": r"\*\*\s*Title:\s*([^*]+?)(?:\*\*|Created|$)",
            "specialty": r"\*\*\s*Specialties:\s*([^*]+?)(?:\*\*|Created|$)",
            "qualifications": r"\*\*\s*Qualifications:\s*([^*]+?)(?:\*\*|Created|$)",
            "department": r"\*\*\s*Department:\s*([^*]+?)(?:\*\*|Created|$)",
        }

        for key, pattern in patterns.items():
            match = re.search(pattern, note_clean, re.IGNORECASE)
            if match:
                result[key] = match.group(1).strip()

        # Also extract HCO name from the note if present (after "HCO_Name:")
        hco_match = re.search(r"\*\*\s*HCO_Name:\s*([^*]+?)(?:\*\*|Created|$)", note_clean, re.IGNORECASE)
        if hco_match:
            result["hco_name"] = hco_match.group(1).strip()

        # Extract additional URLs mentioned after "Created for Inbox" section
        # e.g., "// Degree from https://..." or "| POV - https://..."
        extra_urls = re.findall(r"(?:POV|Degree|pov|degree|source)[^h]*?(https?://\S+)", note_clean)
        if extra_urls:
            result["additional_urls"] = extra_urls

        return result

    @classmethod
    def parse(cls, note, source=None):
        """Auto-detect format and parse."""
        fmt = cls.detect_format(note)
        if fmt == "manual":
            return cls.parse_manual(note)
        elif fmt == "omnia":
            return cls.parse_omnia(note)
        else:
            # Try both and return whichever gets more data
            manual = cls.parse_manual(note)
            omnia = cls.parse_omnia(note)
            manual_hits = sum(1 for v in manual.values() if v and v != "manual")
            omnia_hits = sum(1 for v in omnia.values() if v and v != "omnia")
            if omnia_hits > manual_hits:
                omnia["parse_format"] = "omnia (auto-detected)"
                return omnia
            else:
                manual["parse_format"] = "manual (auto-detected)"
                return manual


# =============================================================================
# VALIDATION CHECKS
# =============================================================================

class Validators:
    """All rule-based validation checks."""

    @staticmethod
    def check_completeness(pdr_row, parsed_note):
        """
        Step 2.1: Check if critical information is present.
        Returns: (is_complete, missing_fields)
        """
        missing = []

        # Check HCP Name
        first_name = str(pdr_row.get("first_name", "")).strip()
        last_name = str(pdr_row.get("last_name", "")).strip()
        hcp_name = parsed_note.get("hcp_name", "")

        if (not first_name or first_name.lower() == "nan") and \
           (not last_name or last_name.lower() == "nan") and \
           (not hcp_name):
            missing.append("HCP Name")

        # Check URL / POV
        hco_url = parsed_note.get("hco_url", "")
        if not hco_url:
            missing.append("Source URL (POV)")

        # Check HCO VID
        hco_vid = parsed_note.get("hco_vid", "")
        if not hco_vid:
            missing.append("HCO VID (Parent Affiliation)")

        return (len(missing) == 0, missing)

    @staticmethod
    def check_scope(parsed_note, pdr_row=None):
        """
        Step 3/3.1: Check if HCP is in scope (MBBS or Business Professional).
        Returns: (is_in_scope, reason)
        """
        # Gather all text that might indicate specialty/degree
        text_sources = []
        for field in ["qualifications", "specialty", "department", "title"]:
            val = parsed_note.get(field, "")
            if val:
                text_sources.append(val)

        combined_text = " ".join(text_sources).lower()

        # If no qualification/specialty info, we can't auto-determine scope
        if not combined_text.strip():
            # Check scope_status from manual format
            scope_status = parsed_note.get("scope_status", "")
            if scope_status and "in scope" in scope_status.lower():
                return (True, "Scope pre-marked as 'In scope' in steward notes")
            return (None, "No qualification/specialty data to assess scope - manual review needed")

        # Check for out-of-scope indicators FIRST
        for oos in Config.OUT_OF_SCOPE_SPECIALTIES:
            if oos.lower() in combined_text:
                return (False, f"Out of scope: '{oos}' detected in qualifications/specialty")

        # Check for in-scope indicators
        for ins in Config.IN_SCOPE_DEGREES:
            if ins.lower() in combined_text:
                return (True, f"In scope: '{ins.upper()}' degree confirmed")

        # Check for business professional titles
        for title in Config.BUSINESS_TITLES:
            if title.lower() in combined_text:
                return (True, f"In scope: Business Professional title '{title}' detected")

        # If we have specialty text but no clear degree match
        return (None, f"Scope uncertain - specialty/dept found but no MBBS/degree confirmed: '{combined_text[:80]}'")

    @staticmethod
    def check_junk_data(pdr_row, parsed_note):
        """
        Detect junk/malformed entries.
        Returns: (is_clean, issues)
        """
        issues = []

        first_name = str(pdr_row.get("first_name", "")).strip()
        last_name = str(pdr_row.get("last_name", "")).strip()
        hcp_name = parsed_note.get("hcp_name", "")
        steward_note = str(pdr_row.get("steward_note", ""))

        # Check for multiple names crammed in one field
        if first_name and ("," in first_name or " and " in first_name.lower() or ";" in first_name):
            issues.append(f"Multiple names detected in first_name: '{first_name}'")

        # Check name length
        full_name = f"{first_name} {last_name}".strip()
        if len(full_name) < Config.MIN_NAME_LENGTH:
            issues.append(f"Name too short: '{full_name}'")
        if len(full_name) > Config.MAX_NAME_LENGTH:
            issues.append(f"Name too long ({len(full_name)} chars): possibly scraped text")

        # Check for junk patterns in steward notes
        for pattern in Config.JUNK_PATTERNS:
            if re.search(pattern, steward_note):
                match_text = re.search(pattern, steward_note).group(0)
                # Only flag if it's in the name fields, not just in the URL/notes
                if re.search(pattern, first_name + " " + last_name + " " + (hcp_name or "")):
                    issues.append(f"Junk pattern detected: '{match_text}'")

        # Check for entity names that aren't individuals
        entity_patterns = [
            r"(?i)^(the\s+)?(association|society|council|federation|forum|committee|trust|foundation)\b",
        ]
        for pattern in entity_patterns:
            if re.search(pattern, full_name) or (hcp_name and re.search(pattern, hcp_name)):
                issues.append(f"Non-individual entity detected in name")

        return (len(issues) == 0, issues)


# =============================================================================
# FUZZY MATCHING ENGINE
# =============================================================================

class FuzzyMatcher:
    """Fuzzy matching against NWK reference data with blocking for performance."""

    def __init__(self, nwk_df):
        """Initialize with NWK reference DataFrame and build index."""
        self.nwk_df = nwk_df
        self.index = defaultdict(list)
        self._build_index()

    def _normalize(self, text):
        """Normalize text for matching."""
        if not text or not isinstance(text, str):
            return ""
        # Remove titles, punctuation, extra spaces
        text = re.sub(r"(?i)^(dr\.?|prof\.?|mr\.?|mrs\.?|ms\.?)\s*", "", text)
        text = re.sub(r"[.\-',()]", " ", text)
        text = re.sub(r"\s+", " ", text).strip().lower()
        return text

    def _get_block_keys(self, first_name, last_name, city):
        """Generate blocking keys for a record."""
        keys = []
        fn = self._normalize(first_name)
        ln = self._normalize(last_name)
        c = self._normalize(city)

        # Block by last_name prefix (first 3 chars) + city
        if ln and len(ln) >= 2:
            ln_prefix = ln[:3]
            if c:
                keys.append(f"{ln_prefix}_{c}")
            keys.append(f"ln_{ln_prefix}")

        # Block by first_name prefix + city (catches last_name variations)
        if fn and len(fn) >= 2:
            fn_prefix = fn[:3]
            if c:
                keys.append(f"{fn_prefix}_{c}")

        # Block by full name soundex-like (first 2 chars of each)
        if fn and ln:
            keys.append(f"fl_{fn[:2]}_{ln[:2]}")

        return keys

    def _build_index(self):
        """Build blocking index from NWK data."""
        print(f"  Building search index for {len(self.nwk_df)} NWK records...")
        for idx, row in self.nwk_df.iterrows():
            fn = str(row.get("first_name", ""))
            ln = str(row.get("last_name", ""))
            city = str(row.get("city", ""))

            keys = self._get_block_keys(fn, ln, city)
            for key in keys:
                self.index[key].append(idx)

        print(f"  Index built: {len(self.index)} blocks, avg {sum(len(v) for v in self.index.values()) / max(len(self.index), 1):.1f} records/block")

    def _similarity(self, s1, s2):
        """Calculate similarity ratio between two strings (0-100)."""
        if not s1 or not s2:
            return 0
        s1 = self._normalize(s1)
        s2 = self._normalize(s2)
        if not s1 or not s2:
            return 0
        return int(SequenceMatcher(None, s1, s2).ratio() * 100)

    def _composite_score(self, pdr_name, pdr_specialty, pdr_city, pdr_affiliation,
                          nwk_row):
        """
        Calculate composite match score.
        Weighted: Name (50%) + Specialty (20%) + City (15%) + Affiliation (15%)
        """
        nwk_name = f"{nwk_row.get('first_name', '')} {nwk_row.get('last_name', '')}".strip()
        nwk_specialty = str(nwk_row.get("specialty_1", ""))
        nwk_city = str(nwk_row.get("city", ""))
        nwk_affiliation = str(nwk_row.get("parent_hco_name", ""))

        name_score = self._similarity(pdr_name, nwk_name)
        spec_score = self._similarity(pdr_specialty, nwk_specialty) if pdr_specialty else 50  # neutral if unknown
        city_score = self._similarity(pdr_city, nwk_city) if pdr_city else 50
        affil_score = self._similarity(pdr_affiliation, nwk_affiliation) if pdr_affiliation else 50

        # Weighted composite
        composite = (name_score * 0.50) + (spec_score * 0.20) + (city_score * 0.15) + (affil_score * 0.15)

        return {
            "composite": int(composite),
            "name_score": name_score,
            "specialty_score": spec_score,
            "city_score": city_score,
            "affiliation_score": affil_score,
            "nwk_name": nwk_name,
            "nwk_specialty": nwk_specialty,
            "nwk_city": nwk_city,
            "nwk_affiliation": nwk_affiliation,
            "nwk_vid": str(nwk_row.get("vid", "")),
        }

    def find_matches(self, first_name, last_name, specialty, city, affiliation, top_n=3):
        """
        Find best matches for a PDR record in NWK.
        Returns list of match dicts sorted by composite score.
        """
        pdr_name = f"{first_name} {last_name}".strip()

        # Get candidate indices from blocking
        block_keys = self._get_block_keys(first_name, last_name, city)
        candidate_indices = set()
        for key in block_keys:
            candidate_indices.update(self.index.get(key, []))

        if not candidate_indices:
            return []

        # Score all candidates
        scores = []
        for idx in candidate_indices:
            nwk_row = self.nwk_df.loc[idx]
            score = self._composite_score(pdr_name, specialty, city, affiliation, nwk_row)
            score["nwk_index"] = idx
            scores.append(score)

        # Sort by composite score descending
        scores.sort(key=lambda x: x["composite"], reverse=True)

        return scores[:top_n]


# =============================================================================
# MAIN PROCESSOR
# =============================================================================

class PDRProcessor:
    """Main processing pipeline."""

    # Column mapping for PDR file
    PDR_COLUMN_MAP = {
        "hcp.vid__v (VID)": "vid",
        "hcp.first_name__v (FIRST NAME)": "first_name",
        "hcp.last_name__v (LAST NAME)": "last_name",
        "hcp.hcp_type__v (HCP TYPE)": "hcp_type",
        "change_request.change_request_id (CHANGE REQUEST ID)": "change_request_id",
        "change_request.source (SOURCE)": "source",
        "change_request_hcp.custom_comments__c_req (STEWARD NOTE (REQUESTED))": "steward_note",
        "change_request_hco.corporate_name__v_req (CORPORATE NAME (REQUESTED))": "corporate_name",
        "change_request.change_request_type (TYPE)": "type",
        "change_request.owner (OWNER)": "owner",
        "change_request.state_key (STATE KEY)": "state_key",
    }

    # Column mapping for NWK reference file
    NWK_COLUMN_MAP = {
        "hcp.vid__v (VID)": "vid",
        "hcp.first_name__v (FIRST NAME)": "first_name",
        "hcp.last_name__v (LAST NAME)": "last_name",
        "hcp.primary_country__v (PRIMARY COUNTRY)": "country",
        "hcp.specialty_1__v (SPECIALTY 1)": "specialty_1",
        "hcp.specialty_2__v (SPECIALTY 2)": "specialty_2",
        "hcp.hcp_status__v (STATUS)": "status",
        "hcp.source_full_name__v (SOURCE FULL NAME)": "source_full_name",
        "hcp.city_cda__v (CITY (CDA))": "city",
        "hcp.custom_degree_1__c (CUSTOM DEGREE 1)": "degree_1",
        "hcp.custom_degree_2__c (CUSTOM DEGREE 2)": "degree_2",
        "hcp.hcp_type__v (HCP TYPE)": "hcp_type",
        "hcp.gender__v (GENDER)": "gender",
        "hco.parent_hco_vid__v (PARENT_HCO_VID__V)": "parent_hco_vid",
        "hco.parent_hco_name__v (PARENT_HCO_NAME__V)": "parent_hco_name",
        "address.locality__v (CITY)": "address_city",
        "address.administrative_area__v (STATE/PROVINCE)": "state",
        "hcp.custom_comments__c (STEWARD NOTE)": "steward_note_nwk",
    }

    def __init__(self, pdr_path=None, nwk_path=None, pdr_df=None, nwk_df=None):
        """
        Initialize processor.
        Can pass file paths OR pre-loaded DataFrames (useful for testing).
        """
        self.pdr_path = pdr_path
        self.nwk_path = nwk_path
        self.pdr_df = pdr_df
        self.nwk_df = nwk_df
        self.results = []
        self.stats = {
            "total": 0,
            "tier1_auto": 0,
            "tier2_assisted": 0,
            "tier3_human": 0,
            "r_10001": 0,
            "r_10058": 0,
            "r_10075": 0,
            "r_10056": 0,
            "a_10017": 0,
            "review": 0,
        }

    def load_data(self):
        """Load and normalize input files."""
        print("\n[1/6] Loading data files...")

        # Load PDR data
        if self.pdr_df is None:
            print(f"  Reading PDR file: {self.pdr_path}")
            self.pdr_df = pd.read_excel(self.pdr_path)
        print(f"  PDR records loaded: {len(self.pdr_df)}")

        # Rename columns to standard names
        rename_map = {}
        for original, standard in self.PDR_COLUMN_MAP.items():
            if original in self.pdr_df.columns:
                rename_map[original] = standard
        self.pdr_df.rename(columns=rename_map, inplace=True)

        # Load NWK reference data
        if self.nwk_df is None:
            if self.nwk_path:
                print(f"  Reading NWK file: {self.nwk_path}")
                self.nwk_df = pd.read_excel(self.nwk_path)
            else:
                print("  WARNING: No NWK reference file provided. Skipping duplicate matching.")
                self.nwk_df = pd.DataFrame()
        print(f"  NWK records loaded: {len(self.nwk_df)}")

        # Rename NWK columns
        if not self.nwk_df.empty:
            rename_map = {}
            for original, standard in self.NWK_COLUMN_MAP.items():
                if original in self.nwk_df.columns:
                    rename_map[original] = standard
            self.nwk_df.rename(columns=rename_map, inplace=True)

    def parse_notes(self):
        """Parse all steward notes."""
        print("\n[2/6] Parsing steward notes...")
        self.parsed_notes = []
        format_counts = defaultdict(int)

        for idx, row in self.pdr_df.iterrows():
            note = row.get("steward_note", "")
            source = row.get("source", "")
            parsed = NoteParser.parse(note, source)
            self.parsed_notes.append(parsed)
            format_counts[parsed.get("parse_format", "unknown")] += 1

        for fmt, count in format_counts.items():
            print(f"  {fmt}: {count} records")

    def run_gate_completeness(self):
        """Gate 1: Completeness check (R-10001)."""
        print("\n[3/6] Running completeness check...")
        flagged = 0

        for i, (idx, row) in enumerate(self.pdr_df.iterrows()):
            parsed = self.parsed_notes[i]
            is_complete, missing = Validators.check_completeness(row, parsed)

            self.results[i]["completeness_pass"] = is_complete
            self.results[i]["missing_fields"] = ", ".join(missing) if missing else ""

            if not is_complete:
                self.results[i]["ai_resolution_code"] = "R-10001"
                self.results[i]["ai_confidence"] = 95
                self.results[i]["ai_tier"] = "Tier 1 (Auto)"
                self.results[i]["recommendation_notes"] = f"Incomplete: missing {', '.join(missing)}"
                self.results[i]["steward_action"] = "SPOT_CHECK"
                self.stats["r_10001"] += 1
                self.stats["tier1_auto"] += 1
                flagged += 1

        print(f"  R-10001 (Incomplete): {flagged} records")

    def run_gate_scope(self):
        """Gate 2: Scope check (R-10058)."""
        print("\n[4/6] Running scope check...")
        flagged = 0
        uncertain = 0

        for i, (idx, row) in enumerate(self.pdr_df.iterrows()):
            # Skip already resolved
            if self.results[i]["ai_resolution_code"]:
                continue

            parsed = self.parsed_notes[i]
            is_in_scope, reason = Validators.check_scope(parsed, row)

            self.results[i]["scope_check"] = "In Scope" if is_in_scope else ("Out of Scope" if is_in_scope is False else "Uncertain")
            self.results[i]["scope_reason"] = reason

            if is_in_scope is False:
                # Extract the out-of-scope specialty for the resolution note
                spec_match = re.search(r"'([^']+)'", reason)
                spec_label = spec_match.group(1) if spec_match else "non-MBBS"

                self.results[i]["ai_resolution_code"] = "R-10058"
                self.results[i]["ai_confidence"] = 90
                self.results[i]["ai_tier"] = "Tier 1 (Auto)"
                self.results[i]["recommendation_notes"] = f"Out of scope: {spec_label} detected"
                self.results[i]["steward_action"] = "SPOT_CHECK"
                self.stats["r_10058"] += 1
                self.stats["tier1_auto"] += 1
                flagged += 1
            elif is_in_scope is None:
                uncertain += 1

        print(f"  R-10058 (Out of Scope): {flagged} records")
        print(f"  Scope uncertain (needs review): {uncertain} records")

    def run_gate_junk(self):
        """Gate 3: Junk data detection (R-10056)."""
        print("\n[4b/6] Running junk data detection...")
        flagged = 0

        for i, (idx, row) in enumerate(self.pdr_df.iterrows()):
            if self.results[i]["ai_resolution_code"]:
                continue

            parsed = self.parsed_notes[i]
            is_clean, issues = Validators.check_junk_data(row, parsed)

            if not is_clean:
                self.results[i]["ai_resolution_code"] = "R-10056"
                self.results[i]["ai_confidence"] = 85
                self.results[i]["ai_tier"] = "Tier 1 (Auto)"
                self.results[i]["recommendation_notes"] = f"Junk data: {'; '.join(issues)}"
                self.results[i]["steward_action"] = "SPOT_CHECK"
                self.stats["r_10056"] += 1
                self.stats["tier1_auto"] += 1
                flagged += 1

        print(f"  R-10056 (Junk Data): {flagged} records")

    def run_gate_duplicate(self):
        """Gate 4: Duplicate matching against NWK (R-10075 / partial match review)."""
        print("\n[5/6] Running duplicate matching...")

        if self.nwk_df.empty:
            print("  SKIPPED: No NWK reference data loaded")
            # Mark remaining as needing review
            for i in range(len(self.results)):
                if not self.results[i]["ai_resolution_code"]:
                    self.results[i]["ai_resolution_code"] = "REVIEW"
                    self.results[i]["ai_confidence"] = 50
                    self.results[i]["ai_tier"] = "Tier 2 (AI-Assisted)"
                    self.results[i]["recommendation_notes"] = "No NWK reference data for matching - manual NWK search required"
                    self.results[i]["steward_action"] = "REVIEW_MATCH"
                    self.stats["review"] += 1
                    self.stats["tier2_assisted"] += 1
            return

        matcher = FuzzyMatcher(self.nwk_df)

        exact_matches = 0
        partial_matches = 0
        no_matches = 0

        for i, (idx, row) in enumerate(self.pdr_df.iterrows()):
            if self.results[i]["ai_resolution_code"]:
                continue

            parsed = self.parsed_notes[i]
            first_name = str(row.get("first_name", "")).strip()
            last_name = str(row.get("last_name", "")).strip()
            specialty = parsed.get("specialty", "") or parsed.get("department", "")
            city = parsed.get("city", "") or str(row.get("city", ""))

            # Get HCO name for affiliation matching
            affiliation = parsed.get("hco_name", "") or str(row.get("corporate_name", ""))

            matches = matcher.find_matches(first_name, last_name, specialty, city, affiliation)

            if matches:
                best = matches[0]
                self.results[i]["match_vid"] = best["nwk_vid"]
                self.results[i]["match_score"] = best["composite"]
                self.results[i]["match_details"] = (
                    f"PDR: {first_name} {last_name} | NWK: {best['nwk_name']} | "
                    f"Name:{best['name_score']}% Spec:{best['specialty_score']}% "
                    f"City:{best['city_score']}% Affil:{best['affiliation_score']}%"
                )

                if best["composite"] >= Config.EXACT_MATCH_THRESHOLD:
                    self.results[i]["ai_resolution_code"] = "R-10075"
                    self.results[i]["ai_confidence"] = best["composite"]
                    self.results[i]["ai_tier"] = "Tier 1 (Auto)"
                    self.results[i]["recommendation_notes"] = (
                        f"Exact match found in NWK: {best['nwk_name']} "
                        f"(VID: {best['nwk_vid']}, {best['nwk_specialty']}, {best['nwk_city']})"
                    )
                    self.results[i]["steward_action"] = "SPOT_CHECK"
                    self.stats["r_10075"] += 1
                    self.stats["tier1_auto"] += 1
                    exact_matches += 1

                elif best["composite"] >= Config.PARTIAL_MATCH_THRESHOLD:
                    # Partial match — needs steward decision
                    self.results[i]["ai_resolution_code"] = "REVIEW"
                    self.results[i]["ai_confidence"] = best["composite"]
                    self.results[i]["ai_tier"] = "Tier 2 (AI-Assisted)"

                    # Determine what differs
                    diffs = []
                    if best["name_score"] < 90:
                        diffs.append(f"Name mismatch ({best['name_score']}%)")
                    if best["specialty_score"] < 80:
                        diffs.append(f"Specialty mismatch ({best['specialty_score']}%)")
                    if best["affiliation_score"] < 70:
                        diffs.append(f"Affiliation mismatch ({best['affiliation_score']}%)")

                    self.results[i]["recommendation_notes"] = (
                        f"Partial match found: {best['nwk_name']} (VID: {best['nwk_vid']}). "
                        f"Differences: {', '.join(diffs) if diffs else 'minor variations'}. "
                        f"May need Phone Validation (Step 6) or URL validation."
                    )
                    self.results[i]["steward_action"] = "REVIEW_MATCH"
                    self.stats["review"] += 1
                    self.stats["tier2_assisted"] += 1
                    partial_matches += 1

                else:
                    # Low match — likely new HCP
                    self.results[i]["ai_resolution_code"] = "A-10017"
                    self.results[i]["ai_confidence"] = 75
                    self.results[i]["ai_tier"] = "Tier 2 (AI-Assisted)"
                    self.results[i]["recommendation_notes"] = (
                        f"No strong NWK match found (best: {best['nwk_name']}, "
                        f"score: {best['composite']}%). Likely new HCP — validate via URL and create record."
                    )
                    self.results[i]["steward_action"] = "CREATE_RECORD"
                    self.stats["a_10017"] += 1
                    self.stats["tier2_assisted"] += 1
                    no_matches += 1
            else:
                # No candidates found at all
                self.results[i]["match_vid"] = ""
                self.results[i]["match_score"] = 0
                self.results[i]["match_details"] = "No blocking candidates found"
                self.results[i]["ai_resolution_code"] = "A-10017"
                self.results[i]["ai_confidence"] = 80
                self.results[i]["ai_tier"] = "Tier 2 (AI-Assisted)"
                self.results[i]["recommendation_notes"] = (
                    "No match found in NWK. Likely new HCP — validate via URL and create record."
                )
                self.results[i]["steward_action"] = "CREATE_RECORD"
                self.stats["a_10017"] += 1
                self.stats["tier2_assisted"] += 1
                no_matches += 1

        print(f"  R-10075 (Exact Match): {exact_matches} records")
        print(f"  REVIEW (Partial Match): {partial_matches} records")
        print(f"  A-10017 (New HCP): {no_matches} records")

    def run(self):
        """Execute the full processing pipeline."""
        print("=" * 60)
        print("  AI-AUGMENTED PDR PROCESSOR")
        print(f"  Run started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 60)

        # Load data
        self.load_data()
        self.stats["total"] = len(self.pdr_df)

        # Initialize results
        self.results = [{
            "ai_resolution_code": "",
            "ai_confidence": 0,
            "ai_tier": "",
            "match_vid": "",
            "match_score": 0,
            "match_details": "",
            "recommendation_notes": "",
            "steward_action": "",
            "completeness_pass": True,
            "missing_fields": "",
            "scope_check": "",
            "scope_reason": "",
            "parsed_hcp_name": "",
            "parsed_specialty": "",
            "parsed_qualifications": "",
            "parsed_hco_vid": "",
            "parsed_hco_name": "",
            "parsed_hco_url": "",
            "parsed_city": "",
            "note_format": "",
        } for _ in range(len(self.pdr_df))]

        # Parse steward notes
        self.parse_notes()

        # Store parsed data in results
        for i, parsed in enumerate(self.parsed_notes):
            self.results[i]["parsed_hcp_name"] = parsed.get("hcp_name", "") or ""
            self.results[i]["parsed_specialty"] = parsed.get("specialty", "") or parsed.get("department", "") or ""
            self.results[i]["parsed_qualifications"] = parsed.get("qualifications", "") or ""
            self.results[i]["parsed_hco_vid"] = parsed.get("hco_vid", "") or ""
            self.results[i]["parsed_hco_name"] = parsed.get("hco_name", "") or ""
            self.results[i]["parsed_hco_url"] = parsed.get("hco_url", "") or ""
            self.results[i]["parsed_city"] = parsed.get("city", "") or ""
            self.results[i]["note_format"] = parsed.get("parse_format", "") or ""

        # Run gates sequentially
        self.run_gate_completeness()   # R-10001
        self.run_gate_scope()          # R-10058
        self.run_gate_junk()           # R-10056
        self.run_gate_duplicate()      # R-10075, REVIEW, A-10017

        # Print summary
        self._print_summary()

        return self.results

    def _print_summary(self):
        """Print processing summary."""
        print("\n" + "=" * 60)
        print("  PROCESSING SUMMARY")
        print("=" * 60)
        total = self.stats["total"]
        t1 = self.stats["tier1_auto"]
        t2 = self.stats["tier2_assisted"]
        t3 = self.stats["tier3_human"]

        print(f"\n  Total PDRs processed:  {total}")
        print(f"\n  Tier 1 (Auto-Resolve): {t1} ({t1/max(total,1)*100:.1f}%)")
        print(f"    - R-10001 (Incomplete):   {self.stats['r_10001']}")
        print(f"    - R-10058 (Out of Scope): {self.stats['r_10058']}")
        print(f"    - R-10056 (Junk Data):    {self.stats['r_10056']}")
        print(f"    - R-10075 (Exact Match):  {self.stats['r_10075']}")
        print(f"\n  Tier 2 (AI-Assisted):  {t2} ({t2/max(total,1)*100:.1f}%)")
        print(f"    - A-10017 (New HCP):      {self.stats['a_10017']}")
        print(f"    - REVIEW (Partial Match):  {self.stats['review']}")
        print(f"\n  Tier 3 (Human Required): {t3} ({t3/max(total,1)*100:.1f}%)")

        print(f"\n  Estimated time savings:")
        time_saved_t1 = t1 * 5  # 5 min saved per auto-resolved
        time_saved_t2 = t2 * 4  # 4 min saved per AI-assisted (from 6 to 2)
        total_saved = time_saved_t1 + time_saved_t2
        print(f"    Tier 1: {t1} x 5 min = {time_saved_t1} min saved")
        print(f"    Tier 2: {t2} x 4 min = {time_saved_t2} min saved")
        print(f"    Total:  {total_saved} min ({total_saved/60:.1f} hours)")
        print("=" * 60)

    def export(self, results, output_path):
        """Export results to a formatted Excel file."""
        print(f"\n[6/6] Exporting to {output_path}...")

        # Build output DataFrame
        output_rows = []
        for i, (idx, row) in enumerate(self.pdr_df.iterrows()):
            out = {}

            # Original PDR columns
            for col in self.pdr_df.columns:
                out[col] = row[col]

            # AI-added columns
            r = results[i]
            out["AI_Resolution_Code"] = r["ai_resolution_code"]
            out["AI_Confidence"] = r["ai_confidence"]
            out["AI_Tier"] = r["ai_tier"]
            out["Match_VID"] = r["match_vid"]
            out["Match_Score"] = r["match_score"]
            out["Match_Details"] = r["match_details"]
            out["Recommendation_Notes"] = r["recommendation_notes"]
            out["Steward_Action"] = r["steward_action"]

            # Parsed fields (for verification)
            out["Parsed_HCP_Name"] = r["parsed_hcp_name"]
            out["Parsed_Specialty"] = r["parsed_specialty"]
            out["Parsed_Qualifications"] = r["parsed_qualifications"]
            out["Parsed_HCO_VID"] = r["parsed_hco_vid"]
            out["Parsed_HCO_Name"] = r["parsed_hco_name"]
            out["Parsed_HCO_URL"] = r["parsed_hco_url"]
            out["Parsed_City"] = r["parsed_city"]
            out["Note_Format"] = r["note_format"]

            output_rows.append(out)

        output_df = pd.DataFrame(output_rows)
        output_df.to_excel(output_path, index=False, engine="openpyxl")

        # Apply formatting
        self._format_excel(output_path, output_df)
        print(f"  Exported {len(output_df)} records to {output_path}")

    def _format_excel(self, path, df):
        """Apply conditional formatting to the output Excel."""
        wb = load_workbook(path)
        ws = wb.active

        # Header styling
        header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        ai_header_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")

        # Find AI column start
        ai_cols = ["AI_Resolution_Code", "AI_Confidence", "AI_Tier", "Match_VID",
                    "Match_Score", "Match_Details", "Recommendation_Notes", "Steward_Action",
                    "Parsed_HCP_Name", "Parsed_Specialty", "Parsed_Qualifications",
                    "Parsed_HCO_VID", "Parsed_HCO_Name", "Parsed_HCO_URL", "Parsed_City", "Note_Format"]
        ai_start_col = None

        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=1, column=col_idx).value
            if cell_val in ai_cols:
                if ai_start_col is None:
                    ai_start_col = col_idx
                ws.cell(row=1, column=col_idx).fill = ai_header_fill
                ws.cell(row=1, column=col_idx).font = header_font
            else:
                ws.cell(row=1, column=col_idx).fill = header_fill
                ws.cell(row=1, column=col_idx).font = header_font

            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", wrap_text=True)

        # Conditional fills for AI tier
        tier1_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        tier2_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
        tier3_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

        # Find AI_Tier column index
        tier_col = None
        code_col = None
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val == "AI_Tier":
                tier_col = col_idx
            if val == "AI_Resolution_Code":
                code_col = col_idx

        if tier_col:
            for row_idx in range(2, ws.max_row + 1):
                tier_val = ws.cell(row=row_idx, column=tier_col).value
                if tier_val and "Tier 1" in str(tier_val):
                    fill = tier1_fill
                elif tier_val and "Tier 2" in str(tier_val):
                    fill = tier2_fill
                elif tier_val and "Tier 3" in str(tier_val):
                    fill = tier3_fill
                else:
                    continue

                # Apply fill to all AI columns in this row
                if ai_start_col:
                    for c in range(ai_start_col, ws.max_column + 1):
                        ws.cell(row=row_idx, column=c).fill = fill

        # Auto-fit column widths (approximate)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_text = str(ws.cell(row=1, column=col_idx).value or "")
            max_len = min(len(header_text) + 2, 35)

            for row_idx in range(2, min(ws.max_row + 1, 20)):
                cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, min(len(cell_val), 45))

            ws.column_dimensions[col_letter].width = max_len

        # Freeze panes (freeze header row + original PDR columns)
        if ai_start_col:
            ws.freeze_panes = f"{get_column_letter(ai_start_col)}2"
        else:
            ws.freeze_panes = "A2"

        # Add autofilter
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

        wb.save(path)


# =============================================================================
# CLI INTERFACE
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="AI-Augmented PDR Processing Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # Process with both PDR and NWK reference:
    python pdr_processor.py --pdr pdrs.xlsx --nwk nwk_export.xlsx --output processed.xlsx

    # Process PDRs only (no duplicate matching):
    python pdr_processor.py --pdr pdrs.xlsx --output processed.xlsx

    # Adjust matching thresholds:
    python pdr_processor.py --pdr pdrs.xlsx --nwk nwk.xlsx --output out.xlsx --exact-threshold 95 --partial-threshold 75
        """
    )
    parser.add_argument("--pdr", required=True, help="Path to PDR export Excel file")
    parser.add_argument("--nwk", required=False, help="Path to NWK reference export Excel file")
    parser.add_argument("--output", required=True, help="Path for output Excel file")
    parser.add_argument("--exact-threshold", type=int, default=92,
                        help="Fuzzy match threshold for exact match (default: 92)")
    parser.add_argument("--partial-threshold", type=int, default=70,
                        help="Fuzzy match threshold for partial match (default: 70)")

    args = parser.parse_args()

    # Update config
    Config.EXACT_MATCH_THRESHOLD = args.exact_threshold
    Config.PARTIAL_MATCH_THRESHOLD = args.partial_threshold

    # Run
    processor = PDRProcessor(pdr_path=args.pdr, nwk_path=args.nwk)
    results = processor.run()
    processor.export(results, args.output)

    print("\nDone! Open the output file and sort by AI_Tier to start processing.")


if __name__ == "__main__":
    main()
