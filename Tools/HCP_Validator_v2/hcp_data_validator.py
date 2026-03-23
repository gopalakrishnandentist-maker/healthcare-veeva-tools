"""
HCP Data Validator Tool v4.0 — Streamlit Edition
Author: Built for OpenData India Operations at Veeva Systems
Purpose: Validate HCP License, Candidate, Affiliation & HCP Status data at scale
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime
import os
import io
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import urllib.parse
import re
import time
import warnings
warnings.filterwarnings('ignore', message='Unverified HTTPS request')

# ─── GK.Ai shared theme ──────────────────────────────────────────────────────
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", ".."))
from gkai_theme import inject_gkai_theme, GKAI_PAGE_CONFIG, render_app_header, render_sidebar_nav

# ─── Page Configuration ───────────────────────────────────────────────────────
st.set_page_config(
    **GKAI_PAGE_CONFIG,
    page_title="HCP Data Validator — Veeva OpenData India",
    page_icon="💊",
)
inject_gkai_theme()


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  ANALYSIS ENGINE  (identical logic — all 4 checks)                      ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

_VID_COL_CANDIDATES = [
    'hcp.vid__v (NETWORK ID)',
    'hcp.vid__v (VID)',
]

def _resolve_vid_col(df):
    """Return whichever VID column alias is present in df, or the default."""
    for c in _VID_COL_CANDIDATES:
        if c in df.columns:
            return c
    return _VID_COL_CANDIDATES[0]  # will trigger missing-column error downstream


def check_licenses(df):
    result = {}
    vid_col = _resolve_vid_col(df)
    license_num_col = 'license.license_number__v (LICENSE)'
    license_status_col = 'license.license_status__v (LICENSE STATUS)'
    first_name_col = 'hcp.first_name__v (FIRST NAME)'
    last_name_col = 'hcp.last_name__v (LAST NAME)'

    missing = [c for c in [vid_col, license_status_col] if c not in df.columns]
    if missing:
        result['error'] = f"Required columns not found: {', '.join(missing)}"
        return result

    work_cols = [vid_col, license_status_col]
    if license_num_col in df.columns:
        work_cols.append(license_num_col)
    if first_name_col in df.columns:
        work_cols.append(first_name_col)
    if last_name_col in df.columns:
        work_cols.append(last_name_col)

    df_work = df[work_cols].copy()
    df_work['is_active'] = (df_work[license_status_col] == 'Active').astype(int)

    if license_num_col in df_work.columns:
        has_num = ~(
            df_work[license_num_col].isna()
            | (df_work[license_num_col].astype(str).str.strip() == '')
            | (df_work[license_num_col].astype(str).str.strip() == '0')
        )
        df_work['is_active'] = (df_work['is_active'] & has_num).astype(int)

    vid_counts = df_work.groupby(vid_col)['is_active'].sum().reset_index()
    vid_counts.columns = ['VID', 'Count_of_Active_Licenses']

    if first_name_col in df.columns and last_name_col in df.columns:
        names = df[[vid_col, first_name_col, last_name_col]].drop_duplicates(subset=[vid_col])
        names.columns = ['VID', 'First_Name', 'Last_Name']
        vid_counts = vid_counts.merge(names, on='VID', how='left')

    no_active = vid_counts[vid_counts['Count_of_Active_Licenses'] == 0]
    result['total_vids'] = len(vid_counts)
    result['vids_no_active'] = len(no_active)
    result['vids_with_active'] = len(vid_counts) - len(no_active)
    result['details'] = no_active
    result['all_counts'] = vid_counts.sort_values('Count_of_Active_Licenses')
    return result


def check_candidates(df):
    result = {}
    vid_col = _resolve_vid_col(df)
    candidate_col = 'hcp.candidate_record__v (CANDIDATE RECORD)'
    first_name_col = 'hcp.first_name__v (FIRST NAME)'
    last_name_col = 'hcp.last_name__v (LAST NAME)'
    rejection_col = 'hcp.ap_candidate_rejection_reason__c (CANDIDATE REVIEW RESULT)'

    if vid_col not in df.columns or candidate_col not in df.columns:
        result['error'] = f"Required columns not found: {vid_col}, {candidate_col}"
        return result

    is_candidate = df[candidate_col].astype(str).str.strip().str.upper() == 'TRUE'
    is_candidate = is_candidate | (df[candidate_col].astype(str).str.strip() == '1')
    candidates = df[is_candidate].copy()

    cols_to_show = [vid_col]
    if first_name_col in df.columns:
        cols_to_show.append(first_name_col)
    if last_name_col in df.columns:
        cols_to_show.append(last_name_col)
    cols_to_show.append(candidate_col)
    if rejection_col in df.columns:
        cols_to_show.append(rejection_col)

    candidate_summary = candidates[cols_to_show].drop_duplicates(subset=[vid_col])
    result['total_candidates'] = len(candidate_summary)
    result['details'] = candidate_summary
    return result


def check_affiliations(df):
    result = {}
    vid_col = _resolve_vid_col(df)
    parent_hco_vid_col = 'hco.parent_hco_vid__v (PARENT_HCO_VID__V)'
    parent_status_col = 'hco.parent_hco_status__v (PARENT_HCO_STATUS__V)'
    first_name_col = 'hcp.first_name__v (FIRST NAME)'
    last_name_col = 'hcp.last_name__v (LAST NAME)'

    if vid_col not in df.columns:
        result['error'] = f"Required column not found: {vid_col}"
        return result

    work_cols = [vid_col]
    if parent_hco_vid_col in df.columns:
        work_cols.append(parent_hco_vid_col)
    if parent_status_col in df.columns:
        work_cols.append(parent_status_col)

    df_work = df[work_cols].copy()
    df_work['has_active_affiliation'] = True

    if parent_hco_vid_col in df_work.columns:
        parent_empty = (
            df_work[parent_hco_vid_col].isna()
            | (df_work[parent_hco_vid_col].astype(str).str.strip() == '')
        )
        df_work['has_active_affiliation'] = df_work['has_active_affiliation'] & ~parent_empty

    if parent_status_col in df_work.columns:
        status_inactive = (df_work[parent_status_col] == 'Inactive')
        df_work['has_active_affiliation'] = df_work['has_active_affiliation'] & ~status_inactive

    df_work['active_count'] = df_work['has_active_affiliation'].astype(int)
    vid_counts = df_work.groupby(vid_col)['active_count'].sum().reset_index()
    vid_counts.columns = ['VID', 'Count_of_Active_Affiliations']
    no_active = vid_counts[vid_counts['Count_of_Active_Affiliations'] == 0].copy()

    if first_name_col in df.columns and last_name_col in df.columns:
        names = df[[vid_col, first_name_col, last_name_col]].drop_duplicates(subset=[vid_col])
        names.columns = ['VID', 'First_Name', 'Last_Name']
        no_active = no_active.merge(names, on='VID', how='left')
        vid_counts_full = vid_counts.merge(names, on='VID', how='left')
    else:
        vid_counts_full = vid_counts

    result['total_vids'] = len(vid_counts)
    result['vids_no_active_aff'] = len(no_active)
    result['vids_with_active_aff'] = len(vid_counts) - len(no_active)
    result['details'] = no_active
    result['all_counts'] = vid_counts_full.sort_values('Count_of_Active_Affiliations')
    return result


def check_hcp_status(df):
    result = {}
    vid_col = _resolve_vid_col(df)
    status_col = 'hcp.hcp_status__v (STATUS)'
    first_name_col = 'hcp.first_name__v (FIRST NAME)'
    last_name_col = 'hcp.last_name__v (LAST NAME)'

    missing = [c for c in [vid_col, status_col] if c not in df.columns]
    if missing:
        result['error'] = f"Required columns not found: {', '.join(missing)}"
        return result

    work_cols = [vid_col, status_col]
    if first_name_col in df.columns:
        work_cols.append(first_name_col)
    if last_name_col in df.columns:
        work_cols.append(last_name_col)

    df_work = df[work_cols].copy()
    vid_status = df_work.drop_duplicates(subset=[vid_col])

    non_active = vid_status[
        vid_status[status_col].astype(str).str.strip() != 'Active'
    ].copy()

    rename_map = {vid_col: 'VID', status_col: 'HCP_Status'}
    if first_name_col in non_active.columns:
        rename_map[first_name_col] = 'First_Name'
    if last_name_col in non_active.columns:
        rename_map[last_name_col] = 'Last_Name'
    non_active = non_active.rename(columns=rename_map)

    total_vids = len(vid_status)
    non_active_count = len(non_active)
    result['total_vids'] = total_vids
    result['vids_non_active'] = non_active_count
    result['vids_active'] = total_vids - non_active_count
    result['details'] = non_active
    return result


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  REVALIDATION ENGINE v3 — 11 fixes applied                             ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
      'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')

# ── SPA / image-only domains ─────────────────────────────────────────────
_SPA_DOMAINS = ['google.com/maps', 'maps.google.', 'maps.app.goo.gl',
                'goo.gl/maps', 'lh3.googleusercontent.com', 'googleusercontent.com/p/']

# ── Common sub-pages where hospitals list doctors (FIX 7) ────────────────
_HOSPITAL_SUBPAGES = [
    '/doctors', '/our-doctors', '/doctor', '/our-team', '/team',
    '/consultants', '/specialists', '/medical-team',
]

# ── Keywords to detect doctor listing links on a page (FIX 12) ──────────
_DOCTOR_LINK_KEYWORDS = [
    'doctor', 'doctors', 'our doctors', 'our team', 'team', 'staff',
    'consultants', 'specialists', 'medical team', 'faculty', 'physicians',
    'find a doctor', 'our specialists', 'meet our', 'medical staff',
    'departments', 'department', 'dermatology', 'cardiology', 'orthopedic',
    'pediatric', 'gynaecology', 'gynecology', 'ent', 'ophthalmology',
    'neurology', 'general medicine', 'surgery',
]

# ── English-to-Devanagari transliteration map (FIX 13) ──────────────────
_DEVANAGARI_MAP = {
    'a': '\u0905', 'aa': '\u0906', 'i': '\u0907', 'ee': '\u0908',
    'u': '\u0909', 'oo': '\u090A', 'e': '\u090F', 'ai': '\u0910',
    'o': '\u0913', 'au': '\u0914',
    'ka': '\u0915', 'kha': '\u0916', 'ga': '\u0917', 'gha': '\u0918',
    'cha': '\u091A', 'chha': '\u091B', 'ja': '\u091C', 'jha': '\u091D',
    'tta': '\u091F', 'ttha': '\u0920', 'dda': '\u0921', 'ddha': '\u0922',
    'nna': '\u0923',
    'ta': '\u0924', 'tha': '\u0925', 'da': '\u0926', 'dha': '\u0927',
    'na': '\u0928', 'pa': '\u092A', 'pha': '\u092B', 'ba': '\u092C',
    'bha': '\u092D', 'ma': '\u092E', 'ya': '\u092F', 'ra': '\u0930',
    'la': '\u0932', 'va': '\u0935', 'wa': '\u0935',
    'sha': '\u0936', 'shha': '\u0937', 'sa': '\u0938', 'ha': '\u0939',
    'ksha': '\u0915\u094D\u0937', 'tra': '\u0924\u094D\u0930',
    'gya': '\u091C\u094D\u091E',
}

# ── Common Indian name → Devanagari lookup (direct mappings for accuracy)
_NAME_DEVANAGARI = {
    'dr': '\u0921\u0949\u0915\u094D\u091F\u0930',
    'doctor': '\u0921\u0949\u0915\u094D\u091F\u0930',
    'kumar': '\u0915\u0941\u092E\u093E\u0930',
    'sharma': '\u0936\u0930\u094D\u092E\u093E',
    'singh': '\u0938\u093F\u0902\u0939',
    'gupta': '\u0917\u0941\u092A\u094D\u0924\u093E',
    'patwa': '\u092A\u091F\u0935\u093E',
    'sachdeva': '\u0938\u091A\u0926\u0947\u0935\u093E',
    'patel': '\u092A\u091F\u0947\u0932',
    'joshi': '\u091C\u094B\u0936\u0940',
    'verma': '\u0935\u0930\u094D\u092E\u093E',
    'mishra': '\u092E\u093F\u0936\u094D\u0930\u093E',
    'pandey': '\u092A\u093E\u0902\u0921\u0947',
    'tiwari': '\u0924\u093F\u0935\u093E\u0930\u0940',
    'yadav': '\u092F\u093E\u0926\u0935',
    'dubey': '\u0926\u0941\u092C\u0947',
    'shukla': '\u0936\u0941\u0915\u094D\u0932\u093E',
    'srivastava': '\u0936\u094D\u0930\u0940\u0935\u093E\u0938\u094D\u0924\u0935',
    'tripathi': '\u0924\u094D\u0930\u093F\u092A\u093E\u0920\u0940',
    'chauhan': '\u091A\u094C\u0939\u093E\u0928',
    'agarwal': '\u0905\u0917\u094D\u0930\u0935\u093E\u0932',
    'jain': '\u091C\u0948\u0928',
    'saxena': '\u0938\u0915\u094D\u0938\u0947\u0928\u093E',
    'mehta': '\u092E\u0947\u0939\u0924\u093E',
    'shah': '\u0936\u093E\u0939',
    'khanna': '\u0916\u0928\u094D\u0928\u093E',
    'kapoor': '\u0915\u092A\u0942\u0930',
    'rajput': '\u0930\u093E\u091C\u092A\u0942\u0924',
    'thakur': '\u0920\u093E\u0915\u0941\u0930',
    'nair': '\u0928\u093E\u092F\u0930',
    'iyer': '\u0905\u0940\u092F\u0930',
    'reddy': '\u0930\u0947\u0921\u094D\u0921\u0940',
    'rao': '\u0930\u093E\u0935',
    'pillai': '\u092A\u093F\u0932\u094D\u0932\u0948',
    'menon': '\u092E\u0947\u0928\u0928',
    'nambiar': '\u0928\u092E\u094D\u092C\u093F\u092F\u093E\u0930',
    'kishore': '\u0915\u093F\u0936\u094B\u0930',
    'asha': '\u0906\u0936\u093E',
    'amrita': '\u0905\u092E\u0943\u0924\u093E',
    'sachin': '\u0938\u091A\u093F\u0928',
    'nitish': '\u0928\u093F\u0924\u0940\u0936',
    'ajay': '\u0905\u091C\u092F',
    'vijay': '\u0935\u093F\u091C\u092F',
    'ravi': '\u0930\u0935\u093F',
    'suresh': '\u0938\u0941\u0930\u0947\u0936',
    'ramesh': '\u0930\u092E\u0947\u0936',
    'mahesh': '\u092E\u0939\u0947\u0936',
    'prakash': '\u092A\u094D\u0930\u0915\u093E\u0936',
    'anil': '\u0905\u0928\u093F\u0932',
    'sunil': '\u0938\u0941\u0928\u0940\u0932',
    'deepak': '\u0926\u0940\u092A\u0915',
    'rakesh': '\u0930\u093E\u0915\u0947\u0936',
    'mukesh': '\u092E\u0941\u0915\u0947\u0936',
    'dinesh': '\u0926\u093F\u0928\u0947\u0936',
    'rajesh': '\u0930\u093E\u091C\u0947\u0936',
    'sanjay': '\u0938\u0902\u091C\u092F',
    'vinod': '\u0935\u093F\u0928\u094B\u0926',
    'manoj': '\u092E\u0928\u094B\u091C',
    'ashok': '\u0905\u0936\u094B\u0915',
    'priya': '\u092A\u094D\u0930\u093F\u092F\u093E',
    'neha': '\u0928\u0947\u0939\u093E',
    'pooja': '\u092A\u0942\u091C\u093E',
    'sunita': '\u0938\u0941\u0928\u0940\u0924\u093E',
    'anita': '\u0905\u0928\u093F\u0924\u093E',
    'sangam': '\u0938\u0902\u0917\u092E',
    'hospital': '\u0905\u0938\u094D\u092A\u0924\u093E\u0932',
    'clinic': '\u0915\u094D\u0932\u093F\u0928\u093F\u0915',
    'chandvania': '\u091A\u0902\u0926\u0935\u093E\u0928\u093F\u092F\u093E',
    'singhal': '\u0938\u093F\u0902\u0918\u0932',
    'patra': '\u092A\u093E\u0924\u094D\u0930\u093E',
    'aparesh': '\u0905\u092A\u0930\u0947\u0936',
    'bandyopadhyay': '\u092C\u0928\u094D\u0926\u094D\u092F\u094B\u092A\u093E\u0927\u094D\u092F\u093E\u092F',
    'george': '\u091C\u0949\u0930\u094D\u091C',
    'anne': '\u090F\u0928',
    'mohammed': '\u092E\u094B\u0939\u092E\u094D\u092E\u0926',
    'muhammed': '\u092E\u0941\u0939\u092E\u094D\u092E\u0926',
    'anas': '\u0905\u0928\u0938',
    'naseef': '\u0928\u0938\u0940\u092B',
    'hafsa': '\u0939\u092B\u094D\u0938\u093E',
    'eram': '\u090F\u0930\u092E',
    'muneera': '\u092E\u0941\u0928\u0940\u0930\u093E',
    'azeez': '\u0905\u091C\u093C\u0940\u091C\u093C',
    'afnitha': '\u0905\u092B\u094D\u0928\u093F\u0924\u093E',
    'anuchandra': '\u0905\u0928\u0941\u091A\u0902\u0926\u094D\u0930',
    'vidhya': '\u0935\u093F\u0926\u094D\u092F\u093E',
    'swaroop': '\u0938\u094D\u0935\u0930\u0942\u092A',
    'sathyan': '\u0938\u0924\u094D\u092F\u0928',
    'krishnan': '\u0915\u0943\u0937\u094D\u0923\u0928',
    'kore': '\u0915\u094B\u0930\u0947',
}

# ── Common Indian name transliteration variants (FIX 4) ─────────────────
_NAME_VARIANTS = {
    'muhammad': ['mohammed', 'muhammed', 'mohamed', 'mohammad', 'mohamad', 'md'],
    'mohammed': ['muhammad', 'muhammed', 'mohamed', 'mohammad', 'mohamad', 'md'],
    'muhammed': ['muhammad', 'mohammed', 'mohamed', 'mohammad', 'mohamad', 'md'],
    'mohamed':  ['muhammad', 'mohammed', 'muhammed', 'mohammad', 'mohamad', 'md'],
    'mohammad': ['muhammad', 'mohammed', 'muhammed', 'mohamed', 'mohamad', 'md'],
    'abdul':    ['abdool', 'abdhul', 'abd'],
    'abdur':    ['abdoor', 'abd'],
    'ahmad':    ['ahmed', 'ahamed', 'ahammed'],
    'ahmed':    ['ahmad', 'ahamed', 'ahammed'],
    'ahamed':   ['ahmad', 'ahmed', 'ahammed'],
    'naseef':   ['nasif', 'naseeph', 'nasief'],
    'nasif':    ['naseef', 'naseeph'],
    'rahim':    ['raheem', 'rahime'],
    'raheem':   ['rahim', 'rahime'],
    'rajan':    ['raajan', 'rajn'],
    'ramesh':   ['ramsh', 'rammesh'],
    'suresh':   ['sursh', 'sureesh'],
    'ganesh':   ['ganeesh', 'gansh'],
    'krishna':  ['krishn', 'krshna', 'krishnaa'],
    'gopal':    ['gopaal', 'gopala'],
    'kumar':    ['kumarr', 'kumaar'],
    'prasad':   ['prashad', 'prasaad'],
    'shankar':  ['sankar', 'shankerr', 'shankhar'],
    'venkat':   ['venkt', 'venkatesh'],
    'lakshmi':  ['laxmi', 'lakshmy'],
    'laxmi':    ['lakshmi', 'lakshmy'],
    'shri':     ['sri', 'shree', 'sree'],
    'sri':      ['shri', 'shree', 'sree'],
    'syed':     ['saiyed', 'saiyad', 'sayyed', 'sayyid'],
    'shaikh':   ['sheikh', 'shaik', 'shk'],
    'sheikh':   ['shaikh', 'shaik', 'shk'],
    'iqbal':    ['ikbal', 'eqbal'],
    'amrit':    ['amrith', 'amrita'],
    'amrita':   ['amritha', 'amrit'],
}

# ── Health directory sites (FIX 5 expanded) ──────────────────────────────
_HEALTH_SITES = [
    'practo.com', 'hexahealth.com', 'lybrate.com', 'justdial.com',
    'credihealth.com', 'docplexus.in', 'bajajfinservhealth.in',
    'apollo247.com', 'medifee.com', 'clinicspots.com',
    'facebook.com', 'jsdl.in',
]


# ───────────────────── UTILITY FUNCTIONS ─────────────────────────────────

def _page_text(html, limit=100000):
    """Extract readable text from HTML, stripping noise."""
    soup = BeautifulSoup(html[:limit * 3], 'html.parser')
    for tag in soup(['script', 'style', 'nav', 'footer', 'header', 'noscript', 'iframe']):
        tag.decompose()
    return soup.get_text(separator=' ', strip=True)[:limit]


def _discover_doctor_links(html, base_url):
    """FIX 12: Scan page HTML for links containing doctor/team/staff keywords.
    Returns list of absolute URLs that likely lead to doctor listing pages."""
    try:
        soup = BeautifulSoup(html[:200000], 'html.parser')
        parsed_base = urllib.parse.urlparse(base_url)
        base_domain = f"{parsed_base.scheme}://{parsed_base.netloc}"
        found_links = []
        seen = set()
        for a in soup.find_all('a', href=True):
            href = a['href'].strip()
            link_text = a.get_text(strip=True).lower()
            href_lower = href.lower()
            # Check both the link text AND the href path for doctor keywords
            combined = link_text + ' ' + href_lower
            if any(kw in combined for kw in _DOCTOR_LINK_KEYWORDS):
                # Make absolute URL
                if href.startswith('http'):
                    abs_url = href
                elif href.startswith('/'):
                    abs_url = base_domain + href
                elif href.startswith('#') or href.startswith('mailto:') or href.startswith('tel:'):
                    continue
                else:
                    abs_url = base_domain + '/' + href
                # Only follow links on same domain
                if parsed_base.netloc in urllib.parse.urlparse(abs_url).netloc:
                    if abs_url not in seen:
                        seen.add(abs_url)
                        found_links.append(abs_url)
        return found_links[:15]  # Cap at 15 to avoid excessive crawling
    except Exception:
        return []


def _transliterate_name(name):
    """FIX 13: Transliterate an English name to Devanagari script.
    Uses direct lookup first, then phonetic fallback.
    Returns a list of possible Devanagari renderings."""
    if not name:
        return []
    parts = name.lower().strip().split()
    results = []
    for part in parts:
        # Direct lookup (most reliable)
        if part in _NAME_DEVANAGARI:
            results.append(_NAME_DEVANAGARI[part])
        else:
            # Phonetic transliteration fallback (basic)
            dev = _phonetic_to_devanagari(part)
            if dev:
                results.append(dev)
    return results


def _phonetic_to_devanagari(word):
    """Basic phonetic English-to-Devanagari transliteration."""
    if not word:
        return ''
    result = []
    i = 0
    word = word.lower()
    while i < len(word):
        matched = False
        # Try longest match first (4, 3, 2, 1 chars)
        for length in [4, 3, 2, 1]:
            chunk = word[i:i+length]
            if chunk in _DEVANAGARI_MAP:
                result.append(_DEVANAGARI_MAP[chunk])
                i += length
                matched = True
                break
        if not matched:
            # Skip unrecognized character
            i += 1
    return ''.join(result) if result else ''


def _search_name_in_text_multilingual(first, last, text):
    """FIX 13: Search for name in both English and Devanagari in the text.
    Returns True if name found in any script."""
    if not text:
        return False
    tl = text.lower()
    # English check
    fl = first.lower().strip() if first else ''
    ll = last.lower().strip() if last else ''
    if fl and ll and fl in tl and ll in tl:
        return True
    # Check with name variants (FIX 4)
    if fl and ll:
        first_vars = _get_name_variants(first)
        last_vars = _get_name_variants(last)
        if any(v in tl for v in first_vars) and any(v in tl for v in last_vars):
            return True
    # Devanagari check
    dev_first_list = _transliterate_name(first) if first else []
    dev_last_list = _transliterate_name(last) if last else []
    for df in dev_first_list:
        if df and df in text:
            for dl in dev_last_list:
                if dl and dl in text:
                    return True
            # Even partial Devanagari match (first name only) is useful signal
            return True
    # Single last name in Devanagari
    for dl in dev_last_list:
        if dl and dl in text:
            return True
    return False


def _is_spa_url(url):
    """Detect SPA/image URLs where text extraction won't work."""
    url_lower = url.lower()
    return any(d in url_lower for d in _SPA_DOMAINS)


def _extract_gmaps_entity(url):
    """Extract place/entity name from Google Maps URL."""
    m = re.search(r'/maps/place/([^/@]+)', url)
    if m:
        return urllib.parse.unquote_plus(m.group(1)).replace('+', ' ')
    parsed = urllib.parse.urlparse(url)
    qs = urllib.parse.parse_qs(parsed.query)
    if 'q' in qs:
        return urllib.parse.unquote_plus(qs['q'][0])
    return ''


def _has_non_latin(text, threshold=0.15):
    """FIX 10: Detect if page has significant non-Latin content (Hindi, etc.)."""
    if not text:
        return False
    sample = text[:5000]
    non_latin = sum(1 for c in sample if ord(c) > 0x024F and not c.isspace())
    return (non_latin / max(len(sample), 1)) > threshold


def _hco_core_keywords(hco_name):
    """Extract meaningful core words from HCO name (drop city, generics)."""
    if not hco_name:
        return []
    skip = {'dept', 'of', 'the', 'and', 'for', 'in', 'at', 'a', '-', '/', '&',
            'pvt', 'ltd', 'private', 'limited'}
    base = hco_name.split(' - ')[0].strip() if ' - ' in hco_name else hco_name
    return [p.lower() for p in re.split(r'[\s/\-&]+', base)
            if len(p) > 2 and p.lower() not in skip]


def _domain_matches_hco(url, hco_name):
    """FIX 3 + FIX 9: Check domain for HCO keywords or acronym match."""
    if not url or not hco_name:
        return False
    try:
        domain = urllib.parse.urlparse(url).netloc.lower().replace('www.', '')
        domain_name = domain.split('.')[0]  # e.g. 'neohospitals' from 'neohospitals.co.in'
    except Exception:
        return False
    keywords = _hco_core_keywords(hco_name)

    # Keyword match: any significant keyword in domain
    if any(kw in domain_name for kw in keywords if len(kw) > 3):
        return True

    # FIX 9: Acronym match — mlnmc == M.L.N.M.C. for "Motilal Nehru Medical College"
    base = hco_name.split(' - ')[0].strip() if ' - ' in hco_name else hco_name
    words = [w for w in re.split(r'[\s\-&/]+', base) if len(w) > 1]
    if len(words) >= 3:
        acronym = ''.join(w[0].lower() for w in words)
        if len(acronym) >= 3 and acronym in domain_name:
            return True
    return False


def _get_name_variants(name_part):
    """Return all known spelling variants for a name part."""
    key = name_part.lower().strip()
    variants = {key}
    if key in _NAME_VARIANTS:
        variants.update(_NAME_VARIANTS[key])
    for k, vs in _NAME_VARIANTS.items():
        if key in vs:
            variants.add(k)
            variants.update(vs)
    return variants


def _proximity_check(first, last, text, window=60):
    """FIX 6: Check that first and last name appear NEAR each other, not just
    anywhere on the page. Returns True if found within `window` words."""
    if not first or not last or not text:
        return False
    tl = text.lower()
    fl = first.lower().strip()
    ll = last.lower().strip()
    # Quick check: both must exist somewhere
    if fl not in tl or ll not in tl:
        # Try variants
        first_vars = _get_name_variants(first)
        last_vars = _get_name_variants(last)
        fl_found = next((v for v in first_vars if v in tl), None)
        ll_found = next((v for v in last_vars if v in tl), None)
        if not fl_found or not ll_found:
            return False
        fl = fl_found
        ll = ll_found
    # Find all positions of first name and last name
    words = tl.split()
    first_positions = [i for i, w in enumerate(words) if fl in w]
    last_positions = [i for i, w in enumerate(words) if ll in w]
    # Check if any pair is within the proximity window
    for fp in first_positions:
        for lp in last_positions:
            if abs(fp - lp) <= window:
                return True
    return False


def _name_score(name, text, require_proximity=False, first_name='', last_name=''):
    """Score how well a name appears in text (0-100).
    FIX 6: When require_proximity=True, enforces first+last appearing near each other."""
    if not name or not text:
        return 0
    nl = name.lower().strip()
    tl = text.lower()

    # Direct exact match (full name as substring)
    if nl in tl:
        return 100

    parts = [p for p in nl.split() if len(p) > 2]
    if not parts:
        return 0

    # Exact parts match
    exact_matched = sum(1 for p in parts if p in tl)
    if exact_matched == len(parts):
        # FIX 6: Proximity check — if first+last provided, verify they're near each other
        if require_proximity and first_name and last_name:
            if not _proximity_check(first_name, last_name, text):
                return 40  # Parts exist but far apart → low score (likely different person)
        return 95

    # Variant matching
    variant_matched = 0
    for p in parts:
        if p in tl:
            variant_matched += 1
        else:
            variants = _get_name_variants(p)
            if any(v in tl for v in variants if v != p):
                variant_matched += 1
    if variant_matched == len(parts):
        if require_proximity and first_name and last_name:
            if not _proximity_check(first_name, last_name, text):
                return 40
        return 90

    best = max(exact_matched, variant_matched)
    return int(best / len(parts) * 75)


def _safe_get(url, timeout=15):
    """HTTP GET with retry on timeout (FIX 11)."""
    for attempt in range(2):
        try:
            r = requests.get(url, headers={'User-Agent': _UA},
                             timeout=timeout, allow_redirects=True, verify=False)
            return r
        except requests.exceptions.Timeout:
            if attempt == 0:
                time.sleep(2)
                continue
            raise
    return None


def _try_subpages(base_url, hcp_first, hcp_last, hco_name, specialty=''):
    """FIX 7 + 8 + 12: When main URL fails, try domain root, common sub-pages,
    AND dynamically discovered doctor/team links from the page."""
    try:
        parsed = urllib.parse.urlparse(base_url)
        root = f"{parsed.scheme}://{parsed.netloc}"
    except Exception:
        return None

    full = f"{hcp_first} {hcp_last}".strip()
    seen_urls = set()

    def _check_page(sub_url):
        """Try fetching a sub-page and check for HCP name."""
        if sub_url in seen_urls:
            return None
        seen_urls.add(sub_url)
        try:
            r = _safe_get(sub_url, timeout=10)
            if r and r.status_code == 200:
                text = _page_text(r.text)
                # Try English name match first
                hcp_score = _name_score(full, text, require_proximity=True,
                                        first_name=hcp_first, last_name=hcp_last)
                if hcp_score >= 75:
                    return dict(url=sub_url, hcp_score=hcp_score, text=text,
                                raw_html=r.text)
                # FIX 13: Try Devanagari/multilingual match on non-Latin pages
                if _has_non_latin(text):
                    if _search_name_in_text_multilingual(hcp_first, hcp_last, text):
                        return dict(url=sub_url, hcp_score=80, text=text,
                                    raw_html=r.text,
                                    notes='Name found in local script')
                return dict(url=sub_url, hcp_score=hcp_score, text=text,
                            raw_html=r.text, no_match=True)
        except Exception:
            pass
        return None

    # Phase A: Try domain root + hardcoded sub-pages
    urls_to_try = [root + '/']
    for sp in _HOSPITAL_SUBPAGES:
        urls_to_try.append(root + sp)
    if specialty:
        spec_slug = specialty.lower().replace(' ', '-')
        urls_to_try.append(root + f'/departments/{spec_slug}')
        urls_to_try.append(root + f'/department/{spec_slug}')
        urls_to_try.append(root + f'/departments/view/{spec_slug}')

    root_html = None
    for sub_url in urls_to_try:
        result = _check_page(sub_url)
        if result and not result.get('no_match'):
            return result
        # Save root page HTML for link discovery
        if sub_url == root + '/' and result and result.get('raw_html'):
            root_html = result['raw_html']

    # Phase B (FIX 12): Discover doctor/team/staff links from root page
    if root_html:
        discovered = _discover_doctor_links(root_html, root)
        for disc_url in discovered:
            result = _check_page(disc_url)
            if result and not result.get('no_match'):
                return result

    return None


# ───────────────────── PHASE 1: URL CHECK ────────────────────────────────

def _check_url(url, hcp_first, hcp_last, hco_name, dept_name='',
               specialty='', timeout=15):
    """Phase 1: Check URL + sub-pages for HCP-HCO co-occurrence."""
    out = dict(url_ok=False, hcp_score=0, hco_score=0, dept_score=0,
               status='skipped', notes='', affiliation_confirmed=False)
    if not url or pd.isna(url) or str(url).strip() in ('', 'nan', 'None'):
        out['notes'] = 'No source URL'
        return out
    url = str(url).strip()
    if not url.startswith('http'):
        url = 'https://' + url

    # ── SPA / image URL handling ─────────────────────────────────────────
    if _is_spa_url(url):
        entity = _extract_gmaps_entity(url)
        out['url_ok'] = True
        if entity:
            out['hco_score'] = max(
                _name_score(hco_name, entity),
                85 if _domain_matches_hco(url, hco_name) else 0)
        if out['hco_score'] >= 70:
            out['status'] = 'partial_hco_gmaps'
            out['notes'] = f'Maps/image entity matches HCO: "{entity[:60]}"'
        else:
            out['status'] = 'spa_unreadable'
            out['notes'] = 'Google Maps/image URL — visual content only'
        return out

    # ── Fetch URL ────────────────────────────────────────────────────────
    domain_is_hco = _domain_matches_hco(url, hco_name)
    text = ''
    http_ok = False

    try:
        r = _safe_get(url, timeout=timeout)
        if r and r.status_code == 200:
            http_ok = True
            text = _page_text(r.text)
        elif r:
            out['notes'] = f'HTTP {r.status_code}'
    except requests.exceptions.ConnectionError:
        out['notes'] = 'URL unreachable'
    except requests.exceptions.Timeout:
        out['notes'] = 'URL timed out'
    except Exception as e:
        out['notes'] = str(e)[:80]

    # ── FIX 8: If main URL failed, try domain root + sub-pages ───────────
    subpages_tried = False
    if not http_ok and domain_is_hco:
        subpages_tried = True
        sub = _try_subpages(url, hcp_first, hcp_last, hco_name, specialty)
        if sub:
            http_ok = True
            text = sub['text']
            url = sub['url']
            out['notes'] = f'Found HCP on sub-page: {url}'

    if not http_ok:
        out['status'] = 'dead_url'
        if not out['notes']:
            out['notes'] = 'URL unreachable'
        return out

    # ── Score names ──────────────────────────────────────────────────────
    out['url_ok'] = True
    full = f"{hcp_first} {hcp_last}".strip()

    # FIX 6: Use proximity-aware scoring for HCP
    out['hcp_score'] = _name_score(full, text, require_proximity=True,
                                    first_name=hcp_first, last_name=hcp_last)
    out['hco_score'] = _name_score(hco_name, text) if hco_name else 0
    out['dept_score'] = _name_score(dept_name, text) if dept_name else 0

    # ── FIX 10 + 13: Non-Latin content detection + transliteration ─────
    has_non_latin = _has_non_latin(text)

    # If English matching failed but page has non-Latin content, try Devanagari
    if has_non_latin and out['hcp_score'] < 75:
        if _search_name_in_text_multilingual(hcp_first, hcp_last, text):
            out['hcp_score'] = 80  # High confidence — name found in local script
            out['notes'] = 'HCP name found in local language (Devanagari/regional)'
        # Also try HCO name in Devanagari
        if hco_name and out['hco_score'] < 70:
            hco_parts = hco_name.split()
            hco_dev = _transliterate_name(' '.join(hco_parts[:3]))
            if any(d in text for d in hco_dev if d):
                out['hco_score'] = 75
                if 'local language' not in out.get('notes', ''):
                    out['notes'] = (out.get('notes', '') +
                                    ' | HCO also matched in local script').strip(' |')

    # ── Determine confidence ─────────────────────────────────────────────
    hcp_ok = out['hcp_score'] >= 75
    hco_ok = out['hco_score'] >= 70

    if hcp_ok and (hco_ok or domain_is_hco):
        out['status'] = 'confirmed'
        out['affiliation_confirmed'] = True
        if domain_is_hco and not hco_ok:
            out['hco_score'] = 85
            out['notes'] = out.get('notes', '') or 'HCP on HCO domain (domain-matched)'
        else:
            out['notes'] = out.get('notes', '') or 'HCP + HCO confirmed on source URL'
    elif hcp_ok:
        out['status'] = 'partial_hcp'
        out['affiliation_confirmed'] = False
        out['notes'] = out.get('notes', '') or 'HCP found, HCO not matched'
    elif hco_ok or domain_is_hco:
        out['status'] = 'partial_hco'
        out['affiliation_confirmed'] = False
        out['hco_score'] = max(out['hco_score'], 85 if domain_is_hco else 0)
        out['notes'] = out.get('notes', '') or 'HCO found/domain matched, HCP not on page'
        # FIX 7 + 12: Try sub-pages (hardcoded + discovered links) for HCP
        # Skip if already tried during dead-URL recovery above
        if not subpages_tried:
            sub = _try_subpages(url, hcp_first, hcp_last, hco_name, specialty)
        else:
            sub = None
        if sub and sub.get('hcp_score', 0) >= 75:
            out['hcp_score'] = sub['hcp_score']
            out['status'] = 'confirmed'
            out['affiliation_confirmed'] = True
            sub_notes = sub.get('notes', '')
            out['notes'] = f'HCP found on sub-page: {sub["url"]}'
            if sub_notes:
                out['notes'] += f' ({sub_notes})'
    elif has_non_latin and out['hcp_score'] < 50:
        out['status'] = 'non_latin_content'
        out['affiliation_confirmed'] = False
        out['notes'] = out.get('notes', '') or 'Page has non-English content (Hindi/regional) — manual review'
    else:
        out['status'] = 'not_found'
        out['affiliation_confirmed'] = False
        out['notes'] = out.get('notes', '') or 'URL active but names not found'

    return out


# ───────────────────── PHASE 2: WEB SEARCH ───────────────────────────────

def _ddg_search_html(query, timeout=12):
    """Backend 1: DuckDuckGo HTML (POST method — mimics real form submission)."""
    try:
        r = requests.post(
            'https://html.duckduckgo.com/html/',
            data={'q': query, 'b': ''},
            headers={'User-Agent': _UA,
                     'Referer': 'https://html.duckduckgo.com/',
                     'Content-Type': 'application/x-www-form-urlencoded'},
            timeout=timeout, verify=False)
        if r.status_code != 200:
            return [], f'DDG HTML: HTTP {r.status_code}'
        soup = BeautifulSoup(r.text, 'html.parser')
        results = []
        for link in soup.select('.result__a'):
            href = link.get('href', '')
            # DDG wraps real URLs in a redirect — extract via uddg param
            if 'uddg=' in href:
                qs = urllib.parse.parse_qs(urllib.parse.urlparse(href).query)
                href = qs.get('uddg', [href])[0]
            if not href or 'duckduckgo' in href:
                continue
            parent = link.find_parent('div', class_='result')
            snippet = (parent.get_text() if parent else link.get_text())
            results.append(dict(url=href, snippet=snippet.lower(), title=link.get_text()))
        return results[:8], ''
    except Exception as e:
        return [], f'DDG HTML error: {str(e)[:60]}'


def _ddg_search_lite(query, timeout=12):
    """Backend 2: DuckDuckGo Lite — lighter page, different blocking rules."""
    try:
        r = requests.post(
            'https://lite.duckduckgo.com/lite/',
            data={'q': query},
            headers={'User-Agent': _UA},
            timeout=timeout, verify=False)
        if r.status_code != 200:
            return [], f'DDG Lite: HTTP {r.status_code}'
        soup = BeautifulSoup(r.text, 'html.parser')
        results = []
        # Lite uses table layout — links are in <a> tags inside table rows
        for a in soup.find_all('a', class_='result-link'):
            href = a.get('href', '')
            if not href or 'duckduckgo' in href:
                continue
            # Get the snippet from the next table row
            row = a.find_parent('tr')
            snippet_row = row.find_next_sibling('tr') if row else None
            snippet = snippet_row.get_text() if snippet_row else a.get_text()
            results.append(dict(url=href, snippet=snippet.lower(), title=a.get_text()))
        # Fallback: try generic <a> tags if no .result-link class
        # Only capture links that look like real results (not nav/ads)
        if not results:
            skip_domains = {'duckduckgo.com', 'duck.com', 'about.com'}
            for a in soup.find_all('a', href=True):
                href = a['href']
                if (href.startswith('http')
                        and not any(sd in href for sd in skip_domains)
                        and len(a.get_text(strip=True)) > 10):
                    results.append(dict(url=href, snippet=a.get_text().lower(),
                                        title=a.get_text()))
        return results[:8], ''
    except Exception as e:
        return [], f'DDG Lite error: {str(e)[:60]}'


def _google_search_scrape(query, timeout=12):
    """Backend 3: Google search scrape fallback."""
    try:
        url = f"https://www.google.com/search?q={urllib.parse.quote_plus(query)}&num=8"
        r = requests.get(url, headers={
            'User-Agent': _UA,
            'Accept': 'text/html,application/xhtml+xml',
            'Accept-Language': 'en-US,en;q=0.9',
        }, timeout=timeout, verify=False)
        if r.status_code != 200:
            return [], f'Google: HTTP {r.status_code}'
        soup = BeautifulSoup(r.text, 'html.parser')
        results = []
        # Google result links are typically in <a> tags with /url?q= prefix
        for a in soup.find_all('a', href=True):
            href = a['href']
            if '/url?q=' in href:
                real = urllib.parse.parse_qs(urllib.parse.urlparse(href).query).get('q', [None])[0]
                if real and real.startswith('http') and 'google' not in real:
                    parent = a.find_parent('div')
                    snippet = parent.get_text() if parent else a.get_text()
                    results.append(dict(url=real, snippet=snippet.lower(),
                                        title=a.get_text()))
            elif href.startswith('http') and 'google' not in href and 'gstatic' not in href:
                results.append(dict(url=href, snippet=a.get_text().lower(),
                                    title=a.get_text()))
        return results[:8], ''
    except Exception as e:
        return [], f'Google error: {str(e)[:60]}'


def _web_search(hcp_first, hcp_last, hco_name, specialty='', city=''):
    """Phase 2: Aggressive multi-backend web search. Any URL under the sun
    that links HCP to HCO is accepted — no restrictions on which websites.
    Returns dict with affiliation_confirmed flag indicating whether a URL was
    found that confirms HCP affiliation to HCO."""
    out = dict(found=False, new_url='', suggested_url='', source='',
               confidence='', notes='', search_debug='',
               affiliation_confirmed=False)
    full = f"Dr. {hcp_first} {hcp_last}".strip()
    full_no_dr = f"{hcp_first} {hcp_last}".strip()
    debug_lines = []
    candidates = []  # Collect ALL candidates: (score, url, confidence, notes)

    last_variants = _get_name_variants(hcp_last) if hcp_last else set()
    first_variants = _get_name_variants(hcp_first) if hcp_first else set()
    all_name_variants = last_variants | first_variants
    hco_core = hco_name.split(' - ')[0].strip() if hco_name and ' - ' in hco_name else hco_name
    hco_keywords = _hco_core_keywords(hco_name)

    # ── Build search queries (cast a wide net — any website is valid) ────
    queries = []
    # Priority 1: Exact name + HCO (best signal for affiliation confirmation)
    if hco_core:
        queries.append(f'"{hcp_first} {hcp_last}" "{hco_core}"')       # Exact match both
        queries.append(f'{full} {hco_core}')                           # Dr. First Last HCO
        queries.append(f'{full_no_dr} {hco_core}')                     # First Last HCO
        queries.append(f'"{hcp_last}" "{hco_core}"')                   # "Last" "HCO"
    # Priority 2: With location context
    if hco_core and city:
        queries.append(f'Dr {hcp_last} {hco_core} {city}')
        queries.append(f'{hcp_last} doctor {hco_core} {city}')
    elif hco_core:
        queries.append(f'Dr {hcp_last} {hco_core}')
    # Priority 3: Specialty + HCO
    if specialty and hco_core:
        queries.append(f'{hcp_last} {specialty} {hco_core}')
        queries.append(f'Dr {hcp_first} {hcp_last} {specialty} {hco_core}')
    if specialty and city:
        queries.append(f'{full} {specialty} {city}')
    # Priority 4: Generic name searches (any website — practo, linkedin, etc.)
    queries.append(f'{full} doctor')
    if city:
        queries.append(f'{full} doctor {city}')
    # Priority 5: Health directory and social media (no site: restriction)
    if hco_core:
        queries.append(f'{hcp_last} {hco_core} practo')
        queries.append(f'{hcp_last} {hco_core} justdial')
        queries.append(f'{hcp_last} {hco_core} lybrate')
    # Priority 6: Broader searches — LinkedIn, Facebook, hospital websites
    if hco_core:
        queries.append(f'Dr {hcp_first} {hcp_last} {hco_core} linkedin')
        queries.append(f'{hcp_first} {hcp_last} {hco_core} facebook')

    debug_lines.append(f'Queries: {len(queries)} | HCO kw: {hco_keywords}')

    # ── Search backends ──────────────────────────────────────────────────
    backends = [
        ('DDG_HTML', _ddg_search_html),
        ('DDG_Lite', _ddg_search_lite),
        ('Google', _google_search_scrape),
    ]

    def _score_candidate(href, snippet):
        """Score a search result. Returns (score, confidence, notes) or None.
        Score: 100=perfect, 80=strong, 60=medium, 40=weak, 20=lead."""
        netloc = urllib.parse.urlparse(href).netloc
        snip_lower = snippet.lower() if snippet else ''

        hcp_in_snip = any(v in snip_lower for v in last_variants) if last_variants else False
        hco_in_snip = any(kw in snip_lower for kw in hco_keywords) if hco_keywords else False
        dm = _domain_matches_hco(href, hco_name)

        # ── Tier 1: Both HCP + HCO in snippet → High confidence ────────
        if hcp_in_snip and (hco_in_snip or dm):
            return (90, 'High', f'HCP+HCO in snippet on {netloc}')

        # ── Try fetching the page for deeper analysis ───────────────────
        try:
            pr = _safe_get(href, timeout=10)
            if pr and pr.status_code == 200:
                ptxt = _page_text(pr.text)
                hcp_s = _name_score(full_no_dr, ptxt, require_proximity=True,
                                    first_name=hcp_first, last_name=hcp_last)
                hco_s = _name_score(hco_name, ptxt) if hco_name else 0

                # Tier 1: Full confirmation on page
                if hcp_s >= 75 and (hco_s >= 70 or dm):
                    return (100, 'High', f'HCP({hcp_s})+HCO({hco_s}) confirmed on {netloc}')

                # Tier 2: HCP found, HCO in domain
                if hcp_s >= 75 and dm:
                    return (85, 'High', f'HCP({hcp_s}) on HCO domain {netloc}')

                # Tier 3: HCP found, no HCO match but it's a health site
                if hcp_s >= 75:
                    return (70, 'Medium', f'HCP({hcp_s}) found on {netloc}')

                # Tier 4: Last name found + HCO match
                last_in = any(v in ptxt.lower() for v in last_variants) if last_variants else False
                if last_in and (hco_s >= 60 or dm):
                    return (55, 'Medium', f'Last name+HCO on {netloc}')

                # Tier 5: Multilingual match
                if _has_non_latin(ptxt):
                    if _search_name_in_text_multilingual(hcp_first, hcp_last, ptxt):
                        return (65, 'Medium', f'Name in local script on {netloc}')

                # Tier 6: Just last name on page
                if last_in:
                    return (35, 'Low', f'Last name on {netloc}, needs manual check')
        except Exception:
            pass

        # ── Tier 7: Snippet-only evidence (no page fetch or fetch failed) ─
        if hcp_in_snip and hco_in_snip:
            return (50, 'Medium', f'HCP+HCO in search snippet ({netloc})')
        if hcp_in_snip:
            return (30, 'Low', f'HCP in search snippet ({netloc})')
        if hco_in_snip or dm:
            return (20, 'Low', f'HCO-related page ({netloc}), needs manual check')

        return None

    # ── Execute searches (try up to 10 queries, stop early on good match) ─
    seen_urls = set()
    best_so_far = 0  # Track best score across all queries
    max_fetches_per_query = 5  # Limit page fetches per query to avoid excess
    for q in queries[:10]:
        search_results = []
        for bname, bfunc in backends:
            results, err = bfunc(q)
            if err:
                debug_lines.append(f'{bname} [{q[:35]}]: {err}')
            if results:
                search_results = results
                debug_lines.append(f'{bname} [{q[:35]}]: {len(results)} hits')
                break
            time.sleep(0.3)

        if not search_results:
            debug_lines.append(f'No results: {q[:50]}')
            continue

        fetches_this_query = 0
        for res in search_results:
            href = res['url']
            if href in seen_urls:
                continue
            seen_urls.add(href)
            if fetches_this_query >= max_fetches_per_query:
                break
            fetches_this_query += 1
            scored = _score_candidate(href, res.get('snippet', ''))
            if scored:
                score, conf, note = scored
                candidates.append((score, href, conf, note))
                best_so_far = max(best_so_far, score)
                # Early exit on high-confidence match
                if score >= 85:
                    best = candidates[-1]
                    out.update(found=True, new_url=best[1], suggested_url=best[1],
                               source=urllib.parse.urlparse(best[1]).netloc,
                               confidence=best[2],
                               notes=best[3],
                               search_debug=' | '.join(debug_lines),
                               affiliation_confirmed=True)
                    return out

        # If we already have a solid candidate (score >= 70), stop searching
        if best_so_far >= 70:
            debug_lines.append(f'Early stop: best score {best_so_far} >= 70')
            break

        time.sleep(1.0)

    # ── Pick the best candidate ─────────────────────────────────────────
    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        best = candidates[0]
        score, url, conf, note = best

        if score >= 50:
            # Strong enough to call "found"
            # Affiliation is confirmed when both HCP + HCO co-occur (score >= 70)
            aff_conf = score >= 70
            out.update(found=True, new_url=url, suggested_url=url,
                       source=urllib.parse.urlparse(url).netloc,
                       confidence=conf, notes=note,
                       affiliation_confirmed=aff_conf)
        else:
            # Weak lead — still suggest the URL for manual review
            out.update(found=False, suggested_url=url,
                       source=urllib.parse.urlparse(url).netloc,
                       confidence=conf,
                       notes=f'Weak lead: {note}',
                       affiliation_confirmed=False)

        # Add runner-up URLs to debug
        if len(candidates) > 1:
            alt = [f'{c[3]} [{c[1][:50]}]' for c in candidates[1:3]]
            debug_lines.append(f'Alt URLs: {"; ".join(alt)}')
    else:
        out['notes'] = 'No web evidence found'

    out['search_debug'] = ' | '.join(debug_lines)
    return out


def export_reval_excel(df_results, file_name):
    """Export revalidation results to formatted Excel with Suggested URL column."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        confirmed = df_results[df_results['Validation_Status'] == 'Confirmed']
        confirmed_new = df_results[df_results['Validation_Status'] == 'Confirmed (New URL)']
        leads = df_results[df_results['Validation_Status'] == 'Suggested Lead']
        unconfirmed = df_results[df_results['Validation_Status'] == 'Unconfirmed']
        partial = df_results[df_results['Validation_Status'] == 'Partial Match']

        total = max(len(df_results), 1)
        auto_conf = len(confirmed) + len(confirmed_new)
        url_coverage = auto_conf + len(leads)

        # Affiliation confirmation stats
        aff_col = 'Affiliation_Confirmed'
        aff_yes = df_results[df_results[aff_col] == 'Yes'] if aff_col in df_results.columns else pd.DataFrame()
        aff_no = df_results[df_results[aff_col] == 'No'] if aff_col in df_results.columns else df_results
        aff_rate = len(aff_yes) / total * 100

        summary = pd.DataFrame([
            {'Metric': 'REVALIDATION REPORT', 'Value': ''},
            {'Metric': '', 'Value': ''},
            {'Metric': 'File', 'Value': file_name},
            {'Metric': 'Total Records', 'Value': len(df_results)},
            {'Metric': 'Date', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
            {'Metric': '', 'Value': ''},
            {'Metric': '--- AFFILIATION CONFIRMATION ---', 'Value': ''},
            {'Metric': 'Affiliations Confirmed', 'Value': len(aff_yes)},
            {'Metric': 'Affiliations Unconfirmed', 'Value': len(aff_no)},
            {'Metric': 'Affiliation Confirmation Rate',
             'Value': f"{aff_rate:.1f}%"},
            {'Metric': '', 'Value': ''},
            {'Metric': '--- VALIDATION BREAKDOWN ---', 'Value': ''},
            {'Metric': 'Confirmed (Existing URL)', 'Value': len(confirmed)},
            {'Metric': 'Confirmed (New URL Found)', 'Value': len(confirmed_new)},
            {'Metric': 'Suggested Leads (Manual Verify)', 'Value': len(leads)},
            {'Metric': 'Partial Match', 'Value': len(partial)},
            {'Metric': 'Unconfirmed — No URL Found', 'Value': len(unconfirmed)},
            {'Metric': '', 'Value': ''},
            {'Metric': 'Auto-Confirmation Rate',
             'Value': f"{(auto_conf / total * 100):.1f}%"},
            {'Metric': 'URL Coverage (incl. Leads)',
             'Value': f"{(url_coverage / total * 100):.1f}%"},
        ])

        # Record type breakdown (for candidate mode)
        if 'Record_Type' in df_results.columns:
            candidates = df_results[df_results['Record_Type'] == 'Candidate']
            existing = df_results[df_results['Record_Type'] == 'Existing']
            if len(candidates) > 0:
                cand_confirmed = len(candidates[candidates.get('Affiliation_Confirmed', pd.Series()) == 'Yes']) if 'Affiliation_Confirmed' in candidates.columns else 0
                cand_rate = cand_confirmed / max(len(candidates), 1) * 100
                summary = pd.concat([summary, pd.DataFrame([
                    {'Metric': '', 'Value': ''},
                    {'Metric': '--- RECORD TYPE BREAKDOWN ---', 'Value': ''},
                    {'Metric': 'Candidate Records (No URL)', 'Value': len(candidates)},
                    {'Metric': 'Existing Records (Had URL)', 'Value': len(existing)},
                    {'Metric': 'Candidate Confirmation Rate', 'Value': f"{cand_rate:.1f}%"},
                ])], ignore_index=True)

        summary.to_excel(writer, sheet_name='Summary', index=False)
        df_results.to_excel(writer, sheet_name='All Records', index=False)
        if len(aff_yes) > 0:
            aff_yes.to_excel(writer, sheet_name='Affiliation Confirmed', index=False)
        if len(aff_no) > 0:
            aff_no.to_excel(writer, sheet_name='Affiliation Unconfirmed', index=False)
        if len(confirmed) > 0:
            confirmed.to_excel(writer, sheet_name='Confirmed', index=False)
        if len(confirmed_new) > 0:
            confirmed_new.to_excel(writer, sheet_name='New URL Found', index=False)
        if len(leads) > 0:
            leads.to_excel(writer, sheet_name='Suggested Leads', index=False)
        if len(unconfirmed) > 0:
            unconfirmed.to_excel(writer, sheet_name='Manual Review', index=False)

        # Add Candidates sheet if Record_Type column exists
        if 'Record_Type' in df_results.columns:
            candidates_df = df_results[df_results['Record_Type'] == 'Candidate']
            if len(candidates_df) > 0:
                candidates_df.to_excel(writer, sheet_name='Candidates', index=False)

    output.seek(0)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = load_workbook(output)
        hfill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
        hfont = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        for sn in wb.sheetnames:
            ws = wb[sn]
            vid_cols = set()
            for cell in ws[1]:
                if cell.value and 'vid' in str(cell.value).lower():
                    vid_cols.add(cell.column)
            for ci in vid_cols:
                for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                    for cell in row:
                        if cell.value is not None:
                            cell.value = str(cell.value)
                            cell.number_format = '@'
            for cell in ws[1]:
                cell.fill = hfill
                cell.font = hfont
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in ws.columns:
                ml = 0
                cl = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            ml = max(ml, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[cl].width = min(ml + 4, 50)
            ws.freeze_panes = 'A2'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    except Exception:
        output.seek(0)
        return output


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  HELPERS                                                                 ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

def make_donut(good, bad, good_label, bad_label, good_color, bad_color, title):
    total = good + bad
    if total == 0:
        return None
    pct_bad = bad / total * 100
    fig = go.Figure(data=[go.Pie(
        labels=[good_label, bad_label], values=[good, bad], hole=0.7,
        marker=dict(colors=[good_color, bad_color], line=dict(color='white', width=3)),
        textinfo='none',
        hovertemplate='%{label}: %{value:,}<br>%{percent}<extra></extra>',
    )])
    fig.update_layout(
        title=dict(text=title, font=dict(size=15, color='#0f172a', family='Inter'),
                   x=0, xanchor='left'),
        showlegend=True,
        legend=dict(orientation='h', yanchor='top', y=-0.05, xanchor='center', x=0.5,
                    font=dict(size=11, color='#64748b')),
        margin=dict(t=50, b=40, l=10, r=10), height=300,
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
        annotations=[
            dict(text=f'<b>{pct_bad:.1f}%</b>', x=0.5, y=0.55,
                 font=dict(size=26, color=bad_color, family='Inter'), showarrow=False),
            dict(text='issues', x=0.5, y=0.42,
                 font=dict(size=11, color='#94a3b8', family='Inter'), showarrow=False),
        ],
    )
    return fig


def export_to_excel(df, results, file_name):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_rows = [
            {'Metric': 'HCP DATA VALIDATION REPORT', 'Value': ''},
            {'Metric': '', 'Value': ''},
            {'Metric': 'File Analyzed', 'Value': file_name or ''},
            {'Metric': 'Total Records', 'Value': len(df)},
            {'Metric': 'Analysis Date', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
            {'Metric': '', 'Value': ''},
        ]

        for key, label, fields in [
            ('license', 'LICENSE CHECK', [
                ('Total VIDs', 'total_vids'), ('VIDs with Active License', 'vids_with_active'),
                ('VIDs WITHOUT Active License', 'vids_no_active')]),
            ('candidate', 'CANDIDATE CHECK', [('Candidate Records Found', 'total_candidates')]),
            ('affiliation', 'AFFILIATION CHECK', [
                ('Total VIDs', 'total_vids'), ('VIDs with Active Affiliation', 'vids_with_active_aff'),
                ('VIDs WITHOUT Active Affiliation', 'vids_no_active_aff')]),
            ('hcp_status', 'HCP STATUS CHECK', [
                ('Total VIDs', 'total_vids'), ('VIDs with Active Status', 'vids_active'),
                ('VIDs with Non-Active Status', 'vids_non_active')]),
        ]:
            if key in results and 'error' not in results[key]:
                res = results[key]
                summary_rows.append({'Metric': f'\u2500\u2500 {label} \u2500\u2500', 'Value': ''})
                for metric_name, metric_key in fields:
                    summary_rows.append({'Metric': metric_name, 'Value': res.get(metric_key, 0)})
                summary_rows.append({'Metric': '', 'Value': ''})

        pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Summary', index=False)

        if 'license' in results and 'error' not in results['license']:
            if len(results['license']['details']) > 0:
                results['license']['details'].to_excel(writer, sheet_name='License Issues', index=False)
            results['license']['all_counts'].to_excel(writer, sheet_name='All VID License Counts', index=False)
        if 'candidate' in results and 'error' not in results['candidate']:
            if len(results['candidate']['details']) > 0:
                results['candidate']['details'].to_excel(writer, sheet_name='Candidate Records', index=False)
        if 'affiliation' in results and 'error' not in results['affiliation']:
            if len(results['affiliation']['details']) > 0:
                results['affiliation']['details'].to_excel(writer, sheet_name='Affiliation Issues', index=False)
            results['affiliation']['all_counts'].to_excel(writer, sheet_name='All VID Affiliation Counts', index=False)
        if 'hcp_status' in results and 'error' not in results['hcp_status']:
            if len(results['hcp_status']['details']) > 0:
                results['hcp_status']['details'].to_excel(writer, sheet_name='HCP Status Issues', index=False)

    output.seek(0)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = load_workbook(output)
        header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            vid_col_indices = set()
            for cell in ws[1]:
                if cell.value and 'vid' in str(cell.value).lower():
                    vid_col_indices.add(cell.column)
            for col_idx in vid_col_indices:
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is not None:
                            cell.value = str(cell.value)
                            cell.number_format = '@'
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
            ws.freeze_panes = 'A2'
        formatted = io.BytesIO()
        wb.save(formatted)
        formatted.seek(0)
        return formatted
    except Exception:
        output.seek(0)
        return output


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  MODE & SIDEBAR                                                          ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

# Initialise workflow mode
if 'mode' not in st.session_state:
    st.session_state.mode = 'home'

uploaded_file = None   # will be set inside sidebar when mode == 'quality'

with st.sidebar:
    render_sidebar_nav(app_title="HCP Data Validator", subtitle="GK.Ai", version="v4.0")

    # ── Home / Back button ───────────────────────────────────────────────
    if st.session_state.mode != 'home':
        if st.button("\u2190  Back to Home", use_container_width=True):
            st.session_state.mode = 'home'
            st.rerun()
        st.markdown("---")

    # ── Mode-specific navigation ─────────────────────────────────────────
    if st.session_state.mode == 'home':
        st.markdown("""<p style="color:#94a3b8;font-size:0.8rem;text-align:center;margin:0;">
            Select a workflow from the main area</p>""", unsafe_allow_html=True)
        page = '_home'

    elif st.session_state.mode == 'quality':
        st.markdown("""<p style="color:#3b82f6;font-size:0.75rem;font-weight:600;
            letter-spacing:0.05em;margin:0 0 8px 0;">DATA QUALITY CHECKS</p>""",
            unsafe_allow_html=True)
        page = st.radio(
            "NAVIGATION", ["\U0001f4ca Dashboard", "\U0001f4cb License Check",
             "\U0001f464 Candidate Records", "\U0001f3e5 Affiliation Check",
             "\U0001f535 HCP Status Check"],
            label_visibility="collapsed",
        )
        st.markdown("---")
        uploaded_file = st.file_uploader(
            "Upload HCP Data File", type=['xlsx', 'xls', 'csv'],
            help="Upload your Veeva OpenData export (.xlsx, .xls, or .csv)",
        )
        if 'results' in st.session_state and st.session_state.results:
            st.markdown("---")
            excel_data = export_to_excel(
                st.session_state.df, st.session_state.results,
                st.session_state.get('file_name', ''))
            st.download_button(
                label="\U0001f4e5  Export to Excel", data=excel_data,
                file_name=f"HCP_Validation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        if 'df' in st.session_state and st.session_state.df is not None:
            st.markdown(f"""
            <div style="background:#0d2818;border:1px solid #166534;border-radius:8px;
                        padding:10px 14px;margin-top:12px;">
                <span style="color:#34d399;font-size:0.78rem;font-weight:600;">
                    \u2713 {st.session_state.get('file_name', 'File loaded')}</span><br>
                <span style="color:#4a5568;font-size:0.7rem;">
                    {len(st.session_state.df):,} rows \u2022 {len(st.session_state.df.columns)} cols</span>
            </div>
            """, unsafe_allow_html=True)

    elif st.session_state.mode == 'revalidation':
        st.markdown("""<p style="color:#f59e0b;font-size:0.75rem;font-weight:600;
            letter-spacing:0.05em;margin:0 0 8px 0;">HCP REVALIDATION</p>""",
            unsafe_allow_html=True)
        page = '_revalidation'
        st.markdown("""<p style="color:#94a3b8;font-size:0.78rem;margin:0;">
            Upload &amp; configure on the main area.<br>Results appear after processing.</p>""",
            unsafe_allow_html=True)

    st.markdown("""
    <div style="text-align:center;margin-top:20px;">
        <span style="background:#1a2744;color:#4a5568;font-size:0.65rem;padding:3px 10px;
                     border-radius:10px;font-family:monospace;">v4.0</span>
    </div>
    """, unsafe_allow_html=True)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  FILE LOADING                                                            ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

if uploaded_file is not None:
    if 'file_name' not in st.session_state or st.session_state.file_name != uploaded_file.name:
        with st.spinner("Loading file..."):
            try:
                if uploaded_file.name.lower().endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                elif uploaded_file.name.lower().endswith('.xls') and not uploaded_file.name.lower().endswith('.xlsx'):
                    df = pd.read_excel(uploaded_file)
                else:
                    df = pd.read_excel(uploaded_file, engine='openpyxl')

                df.columns = df.columns.str.strip()
                for col in df.columns:
                    if 'vid' in col.lower():
                        df[col] = df[col].apply(
                            lambda x: str(int(x)) if pd.notna(x) and isinstance(x, float) and x == int(x)
                            else (str(x) if pd.notna(x) else ''))

                st.session_state.df = df
                st.session_state.file_name = uploaded_file.name
                st.session_state.results = {}
            except Exception as e:
                st.error(f"Failed to load file: {str(e)}")


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  PAGES                                                                   ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

if page == '_home':
    # ── Hero Banner ──────────────────────────────────────────────────────
    render_app_header(
        title="HCP Data Validator",
        description="Veeva OpenData India — License, Candidate, Affiliation, HCP Status & Revalidation checks",
        tags=[
            {"label": "v4.0", "color": "green"},
            {"label": "License", "color": "blue"},
            {"label": "Candidate", "color": "blue"},
            {"label": "Affiliation", "color": "blue"},
            {"label": "HCP Status", "color": "blue"},
            {"label": "Revalidation", "color": "amber"},
        ],
    )

    st.markdown("### What would you like to do?")
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Two Workflow Pods ────────────────────────────────────────────────
    pod1, pod2 = st.columns(2, gap="large")

    with pod1:
        st.markdown("""
        <div style="background:linear-gradient(135deg,rgba(37,99,235,0.12),rgba(59,130,246,0.05));
                    border:1px solid rgba(59,130,246,0.25);border-radius:16px;
                    padding:32px 28px;min-height:320px;">
            <div style="width:56px;height:56px;border-radius:14px;
                        background:linear-gradient(135deg,#2563eb,#3b82f6);
                        display:inline-flex;align-items:center;justify-content:center;
                        margin-bottom:18px;box-shadow:0 4px 16px rgba(37,99,235,0.3);">
                <span style="color:white;font-size:26px;">\u2611</span>
            </div>
            <h3 style="color:#ffffff;font-weight:800;font-size:1.3rem;margin:0 0 8px;">
                Data Quality Checks</h3>
            <p style="color:#94a3b8;font-size:0.88rem;line-height:1.6;margin:0 0 16px;">
                Run 4 validation checks on your HCP data to identify issues
                before they reach production.</p>
            <div style="color:#60a5fa;font-size:0.8rem;line-height:1.8;">
                \u26bf License Check<br>
                \u2637 Candidate Record Check<br>
                \u2695 Affiliation Check<br>
                \u25cf HCP Status Check
            </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("\u2611  Open Quality Checks", type="primary", use_container_width=True):
            st.session_state.mode = 'quality'
            st.rerun()

    with pod2:
        st.markdown("""
        <div style="background:linear-gradient(135deg,rgba(245,158,11,0.12),rgba(239,68,68,0.05));
                    border:1px solid rgba(245,158,11,0.25);border-radius:16px;
                    padding:32px 28px;min-height:320px;">
            <div style="width:56px;height:56px;border-radius:14px;
                        background:linear-gradient(135deg,#f59e0b,#ef4444);
                        display:inline-flex;align-items:center;justify-content:center;
                        margin-bottom:18px;box-shadow:0 4px 16px rgba(245,158,11,0.3);">
                <span style="color:white;font-size:26px;">\U0001f50d</span>
            </div>
            <h3 style="color:#ffffff;font-weight:800;font-size:1.3rem;margin:0 0 8px;">
                HCP Revalidation</h3>
            <p style="color:#94a3b8;font-size:0.88rem;line-height:1.6;margin:0 0 16px;">
                Validate HCP\u2013HCO affiliations: confirm existing URLs,
                discover new validation sources via web search.</p>
            <div style="color:#fbbf24;font-size:0.8rem;line-height:1.8;">
                \U0001f7e2 Phase 1 \u2014 URL validation & affiliation confirmation<br>
                \U0001f535 Phase 2 \u2014 Google/DuckDuckGo web search (any website)<br>
                \u2705 Clear Affiliation Confirmed/Unconfirmed status<br>
                \U0001f4ca Categorized Excel report with validation URLs
            </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("\U0001f50d  Open Revalidation", type="secondary", use_container_width=True):
            st.session_state.mode = 'revalidation'
            st.rerun()

    # ── Footer info ──────────────────────────────────────────────────────
    st.markdown(f"""<div class="footer-bar">
        <span>\u25cf Built for OpenData India Operations</span>
        <span>Veeva Systems \u2022 v4.0</span>
    </div>""", unsafe_allow_html=True)


elif page == "\U0001f4ca Dashboard":
    st.markdown("## \U0001f4ca Dashboard")

    with st.expander("**Select Validation Checks**", expanded=True):
        c1, c2, c3, c4 = st.columns(4)
        run_license = c1.checkbox("\U0001f535 License Check", value=True)
        run_candidate = c2.checkbox("\U0001f7e0 Candidate Check", value=True)
        run_affiliation = c3.checkbox("\U0001f7e2 Affiliation Check", value=True)
        run_hcp_status = c4.checkbox("\U0001f7e3 HCP Status Check", value=True)

    if st.button("\u25b6  Run Analysis", type="primary", use_container_width=True,
                  disabled=('df' not in st.session_state or st.session_state.get('df') is None)):
        if not (run_license or run_candidate or run_affiliation or run_hcp_status):
            st.warning("Please select at least one validation check!")
        else:
            df = st.session_state.df
            results = {}
            with st.spinner("Running validation checks..."):
                if run_license:
                    results['license'] = check_licenses(df)
                if run_candidate:
                    results['candidate'] = check_candidates(df)
                if run_affiliation:
                    results['affiliation'] = check_affiliations(df)
                if run_hcp_status:
                    results['hcp_status'] = check_hcp_status(df)
            st.session_state.results = results
            st.success("\u2713 Analysis complete!")
            st.rerun()

    if 'results' in st.session_state and st.session_state.results:
        results = st.session_state.results
        df = st.session_state.df

        st.markdown("---")

        total_records = len(df)
        total_vids = license_issues = candidate_count = affiliation_issues = hcp_status_issues = 0

        if 'license' in results and 'error' not in results['license']:
            total_vids = results['license']['total_vids']
            license_issues = results['license']['vids_no_active']
        if 'candidate' in results and 'error' not in results['candidate']:
            candidate_count = results['candidate']['total_candidates']
        if 'affiliation' in results and 'error' not in results['affiliation']:
            if not total_vids: total_vids = results['affiliation']['total_vids']
            affiliation_issues = results['affiliation']['vids_no_active_aff']
        if 'hcp_status' in results and 'error' not in results['hcp_status']:
            if not total_vids: total_vids = results['hcp_status']['total_vids']
            hcp_status_issues = results['hcp_status']['vids_non_active']

        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("Total Records", f"{total_records:,}")
        m2.metric("Unique VIDs", f"{total_vids:,}")
        m3.metric("License Issues", f"{license_issues:,}")
        m4.metric("Candidates", f"{candidate_count:,}")
        m5.metric("Affil. Issues", f"{affiliation_issues:,}")
        m6.metric("Status Issues", f"{hcp_status_issues:,}")

        st.markdown("<br>", unsafe_allow_html=True)

        chart_cols = st.columns(3)
        ci = 0
        if 'license' in results and 'error' not in results['license']:
            r = results['license']
            fig = make_donut(r['vids_with_active'], r['vids_no_active'],
                             'With License', 'Without License', '#3b82f6', '#ef4444', 'License Check')
            if fig: chart_cols[ci].plotly_chart(fig, use_container_width=True); ci += 1
        if 'affiliation' in results and 'error' not in results['affiliation']:
            r = results['affiliation']
            fig = make_donut(r['vids_with_active_aff'], r['vids_no_active_aff'],
                             'With Affiliation', 'Without Affiliation', '#10b981', '#ef4444', 'Affiliation Check')
            if fig: chart_cols[min(ci,2)].plotly_chart(fig, use_container_width=True); ci += 1
        if 'hcp_status' in results and 'error' not in results['hcp_status']:
            r = results['hcp_status']
            fig = make_donut(r['vids_active'], r['vids_non_active'],
                             'Active', 'Non-Active', '#8b5cf6', '#ef4444', 'HCP Status Check')
            if fig: chart_cols[min(ci,2)].plotly_chart(fig, use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("### Validation Summary")

        table_html = '<table class="summary-table"><tr><th>Check</th><th>Total VIDs</th><th>Issues Found</th><th>Issue Rate</th><th>Status</th></tr>'
        for key, name, vid_key, issue_key in [
            ('license', 'License Check', 'total_vids', 'vids_no_active'),
            ('candidate', 'Candidate Check', None, 'total_candidates'),
            ('affiliation', 'Affiliation Check', 'total_vids', 'vids_no_active_aff'),
            ('hcp_status', 'HCP Status Check', 'total_vids', 'vids_non_active'),
        ]:
            if key in results:
                r = results[key]
                if 'error' in r:
                    table_html += f'<tr><td>{name}</td><td colspan="4">ERROR: {r["error"]}</td></tr>'
                else:
                    vids = r.get(vid_key, 0) if vid_key else 0
                    issues = r.get(issue_key, 0)
                    if vid_key and vids > 0:
                        pct = f"{(issues/vids*100):.1f}%"
                        thresh = 10 if key != 'hcp_status' else 5
                        if key == 'affiliation': thresh = 15
                        status = 'PASS' if issues == 0 else ('WARNING' if (issues/vids*100) < thresh else 'CRITICAL')
                    elif key == 'candidate':
                        pct = '\u2014'
                        status = 'PASS' if issues == 0 else 'REVIEW'
                    else:
                        pct = '0%'
                        status = 'PASS'
                    rc = status.lower()
                    pill = f'<span class="status-{rc}">{status}</span>'
                    vids_str = f'{vids:,}' if vid_key else '\u2014'
                    table_html += f'<tr class="{rc}-row"><td>{name}</td><td>{vids_str}</td><td>{issues:,}</td><td>{pct}</td><td>{pill}</td></tr>'
        table_html += '</table>'
        st.markdown(table_html, unsafe_allow_html=True)

        st.markdown(f"""<div class="footer-bar">
            <span>\u25cf Analysis complete</span>
            <span>\u23f0 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} \u2022 {st.session_state.get('file_name','N/A')}</span>
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown("""<div style="text-align:center;padding:60px 20px;background:#ffffff;
            border:1px solid #e8edf2;border-radius:12px;margin-top:20px;">
            <p style="font-size:2.5rem;margin:0;">\u2630</p>
            <p style="color:#94a3b8;font-size:1rem;margin:8px 0 4px;">Load a file and run analysis to see results</p>
            <p style="color:#cbd5e1;font-size:0.85rem;">Upload using the sidebar, then hit Run Analysis</p>
        </div>""", unsafe_allow_html=True)


def render_detail_page(icon, title, color, result_key):
    st.markdown(f"""<div class="detail-header">
        <div class="detail-icon" style="background:{color};">{icon}</div>
        <h2>{title}</h2></div>""", unsafe_allow_html=True)
    st.markdown(f'<div style="height:3px;background:linear-gradient(90deg,{color},transparent);'
                f'border-radius:2px;margin-bottom:20px;"></div>', unsafe_allow_html=True)
    if 'results' not in st.session_state or result_key not in st.session_state.results:
        st.markdown("""<div style="text-align:center;padding:60px 20px;background:#ffffff;
            border:1px solid #e8edf2;border-radius:12px;">
            <p style="font-size:2rem;margin:0;color:#e2e8f0;">\u2630</p>
            <p style="color:#94a3b8;margin:8px 0 0;">Run analysis to view results.</p></div>""",
            unsafe_allow_html=True)
        return None
    res = st.session_state.results[result_key]
    if 'error' in res:
        st.error(f"Error: {res['error']}")
        return None
    return res


if page == "\U0001f4cb License Check":
    res = render_detail_page("\u26bf", "License Check Results", "#3b82f6", "license")
    if res:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total VIDs", f'{res["total_vids"]:,}')
        c2.metric("With License", f'{res["vids_with_active"]:,}')
        c3.metric("Without License", f'{res["vids_no_active"]:,}')
        c4.metric("Issue Rate", f'{(res["vids_no_active"]/max(res["total_vids"],1)*100):.2f}%')
        st.markdown("<br>", unsafe_allow_html=True)
        if len(res['details']) == 0:
            st.markdown("""<div class="all-clear"><div class="icon">\u2713</div>
                <h3>No license issues found. All clear!</h3>
                <p>All VIDs have at least one active license.</p></div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"**\u26a0 {len(res['details']):,} VIDs without active license:**")
            st.dataframe(res['details'], use_container_width=True, height=400)

elif page == "\U0001f464 Candidate Records":
    res = render_detail_page("\u2637", "Candidate Record Results", "#f59e0b", "candidate")
    if res:
        st.metric("Candidate Records Found", f'{res["total_candidates"]:,}')
        st.markdown("<br>", unsafe_allow_html=True)
        if len(res['details']) == 0:
            st.markdown("""<div class="all-clear"><div class="icon">\u2713</div>
                <h3>No candidate records found. All clear!</h3></div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"**\u26a0 {len(res['details']):,} candidate records:**")
            st.dataframe(res['details'], use_container_width=True, height=400)

elif page == "\U0001f3e5 Affiliation Check":
    res = render_detail_page("\u2695", "Affiliation Check Results", "#10b981", "affiliation")
    if res:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total VIDs", f'{res["total_vids"]:,}')
        c2.metric("With Affiliation", f'{res["vids_with_active_aff"]:,}')
        c3.metric("Without Affiliation", f'{res["vids_no_active_aff"]:,}')
        c4.metric("Issue Rate", f'{(res["vids_no_active_aff"]/max(res["total_vids"],1)*100):.2f}%')
        st.markdown("<br>", unsafe_allow_html=True)
        if len(res['details']) == 0:
            st.markdown("""<div class="all-clear"><div class="icon">\u2713</div>
                <h3>No affiliation issues found. All clear!</h3></div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"**\u26a0 {len(res['details']):,} VIDs without active affiliation:**")
            st.dataframe(res['details'], use_container_width=True, height=400)

elif page == "\U0001f535 HCP Status Check":
    res = render_detail_page("\u25cf", "HCP Status Check Results", "#8b5cf6", "hcp_status")
    if res:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total VIDs", f'{res["total_vids"]:,}')
        c2.metric("Active Status", f'{res["vids_active"]:,}')
        c3.metric("Non-Active", f'{res["vids_non_active"]:,}')
        c4.metric("Issue Rate", f'{(res["vids_non_active"]/max(res["total_vids"],1)*100):.2f}%')
        st.markdown("<br>", unsafe_allow_html=True)
        if len(res['details']) == 0:
            st.markdown("""<div class="all-clear"><div class="icon">\u2713</div>
                <h3>No HCP status issues found. All clear!</h3>
                <p>All VIDs have Active status.</p></div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"**\u26a0 {len(res['details']):,} VIDs with non-Active status:**")
            st.dataframe(res['details'], use_container_width=True, height=400)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  REVALIDATION PAGE                                                       ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

elif page == '_revalidation':
    st.markdown("""<div class="detail-header">
        <div class="detail-icon" style="background:linear-gradient(135deg,#f59e0b,#ef4444);">\U0001f50d</div>
        <h2>HCP Revalidation</h2></div>""", unsafe_allow_html=True)
    st.markdown('<div style="height:3px;background:linear-gradient(90deg,#f59e0b,#ef4444,transparent);'
                'border-radius:2px;margin-bottom:20px;"></div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="section-card">
        <h3>HCP-HCO Affiliation Validation & URL Discovery</h3>
        <p style="color:#94a3b8;font-size:0.9rem;line-height:1.7;">
            <strong style="color:#f0f0f0;">Primary Goal:</strong> Validate existing HCP data in NWK &
            identify validation URLs where existing sources do not work.<br><br>
            <span style="color:#34d399;font-weight:600;">Phase 1</span> \u2014
            Validate existing URL and <strong>CONFIRM</strong> if the HCP Affiliation to the HCO is confirmed<br>
            <span style="color:#60a5fa;font-weight:600;">Phase 2</span> \u2014
            When a URL cannot be validated, web search via Google/DuckDuckGo to find <strong>any</strong>
            web link confirming HCP\u2013HCO affiliation (no URL restrictions \u2014 any publicly searchable website)<br>
            <span style="color:#a78bfa;font-weight:600;">Candidate Mode</span> \u2014
            For unvalidated candidate records <strong>without</strong> a Source URL \u2014 skip Phase 1 and go
            straight to Phase 2 web search to discover affiliation evidence<br>
            <span style="color:#fbbf24;font-weight:600;">Output</span> \u2014
            Clear Affiliation Confirmed/Unconfirmed status with validation URLs and categorized Excel report
        </p>
    </div>
    """, unsafe_allow_html=True)

    # ── File Upload ──────────────────────────────────────────────────────
    reval_file = st.file_uploader(
        "Upload Revalidation Batch", type=['xlsx', 'xls', 'csv'],
        help="Upload HCP data with source URLs for revalidation",
        key="reval_uploader",
    )

    if reval_file is not None:
        if ('reval_file_name' not in st.session_state
                or st.session_state.get('reval_file_name') != reval_file.name):
            with st.spinner("Loading revalidation data\u2026"):
                try:
                    if reval_file.name.lower().endswith('.csv'):
                        rdf = pd.read_csv(reval_file)
                    elif reval_file.name.lower().endswith('.xls') and not reval_file.name.lower().endswith('.xlsx'):
                        rdf = pd.read_excel(reval_file)
                    else:
                        rdf = pd.read_excel(reval_file, engine='openpyxl')
                    rdf.columns = rdf.columns.str.strip()
                    for col in rdf.columns:
                        if 'vid' in col.lower():
                            rdf[col] = rdf[col].apply(
                                lambda x: str(int(x)) if pd.notna(x) and isinstance(x, float) and x == int(x)
                                else (str(x) if pd.notna(x) else ''))
                    st.session_state.reval_df = rdf
                    st.session_state.reval_file_name = reval_file.name
                    st.session_state.reval_results = None
                except Exception as e:
                    st.error(f"Failed to load file: {e}")

    if 'reval_df' in st.session_state and st.session_state.reval_df is not None:
        rdf = st.session_state.reval_df
        st.success(f"Loaded **{len(rdf):,}** records \u2022 {len(rdf.columns)} columns")

        # ── Processing Options ───────────────────────────────────────────
        st.markdown("### Processing Options")
        oc1, oc2, oc3, oc4 = st.columns(4)
        with oc1:
            candidate_mode = st.checkbox(
                "\U0001f7e3 Candidate Validation (No URL)",
                value=False,
                help="Enable for candidate records without Source URLs. "
                     "Records without a URL skip Phase 1 and go directly to "
                     "Phase 2 web search. Records WITH a URL still go through Phase 1.",
                key="reval_candidate_mode")
        with oc2:
            run_phase2 = st.checkbox("Enable Phase 2 (Web Search)", value=True,
                                     help="Search the web when existing URL fails. Slower but higher confirmation rate.",
                                     disabled=candidate_mode)
        with oc3:
            max_workers = st.slider("Concurrent Connections", 5, 20, 10,
                                    help="Higher = faster but may trigger rate limits")
        with oc4:
            max_records = st.number_input("Max Records (0 = all)", min_value=0, value=0, step=100,
                                          help="Process only first N records for testing")

        # Auto-enable Phase 2 when candidate mode is active
        if candidate_mode:
            run_phase2 = True
            st.info(
                "**Candidate Mode active:** URL column is optional. "
                "Records without a URL will skip Phase 1 and go directly to Phase 2 web search. "
                "Phase 2 is auto-enabled.")

        # ── Column Mapping ───────────────────────────────────────────────
        st.markdown("### Column Configuration")
        all_cols = ['\u2014 Not Available \u2014'] + list(rdf.columns)

        def _auto(keywords):
            for kw in keywords:
                for c in rdf.columns:
                    if kw.lower() in c.lower():
                        return c
            return None

        url_def = _auto(['source_url', 'SOURCE URL', 'validation_source_url', 'url'])
        hco_def = _auto(['grandparent_hco_name', 'GRANDPARENT_HCO_NAME'])
        dept_def = _auto(['parent_hco_name', 'PARENT_HCO_NAME'])
        fn_def = _auto(['first_name', 'FIRST NAME'])
        ln_def = _auto(['last_name', 'LAST NAME'])
        vid_def = _auto(['vid__v', 'VID'])
        spec_def = _auto(['specialty', 'SPECIALTY'])
        city_def = _auto(['city', 'CITY'])

        c1, c2 = st.columns(2)
        with c1:
            url_label = "Source URL Column" + (" (optional)" if candidate_mode else " *")
            url_col = st.selectbox(url_label, all_cols,
                                   index=all_cols.index(url_def) if url_def else 0)
            hco_col = st.selectbox("HCO Name (Grandparent) *", all_cols,
                                   index=all_cols.index(hco_def) if hco_def else 0,
                                   help="hco.grandparent_hco_name__v \u2014 the actual hospital name")
            dept_col = st.selectbox("Department (Parent HCO)", all_cols,
                                    index=all_cols.index(dept_def) if dept_def else 0,
                                    help="hco.parent_hco_name__v \u2014 department-level affiliation (optional)")
            vid_col_r = st.selectbox("VID Column", all_cols,
                                     index=all_cols.index(vid_def) if vid_def else 0,
                                     key="reval_vid")
        with c2:
            fn_col = st.selectbox("HCP First Name *", all_cols,
                                  index=all_cols.index(fn_def) if fn_def else 0, key="reval_fn")
            ln_col = st.selectbox("HCP Last Name *", all_cols,
                                  index=all_cols.index(ln_def) if ln_def else 0, key="reval_ln")
            spec_col = st.selectbox("Specialty (optional)", all_cols,
                                    index=all_cols.index(spec_def) if spec_def else 0, key="reval_spec")
            city_col = st.selectbox("City (optional)", all_cols,
                                    index=all_cols.index(city_def) if city_def else 0, key="reval_city")

        na = '\u2014 Not Available \u2014'
        if candidate_mode:
            required_ok = all(c != na for c in [hco_col, fn_col, ln_col])
        else:
            required_ok = all(c != na for c in [url_col, hco_col, fn_col, ln_col])

        if not required_ok:
            if candidate_mode:
                st.warning("Please map HCO Name, First Name, and Last Name columns to proceed.")
            else:
                st.warning("Please map all required columns (*) to proceed.")

        # ── Run Button ───────────────────────────────────────────────────
        if st.button("\U0001f680 Start Revalidation", type="primary",
                     use_container_width=True, disabled=not required_ok):
            dept_c = dept_col if dept_col != na else None
            spec_c = spec_col if spec_col != na else None
            city_c = city_col if city_col != na else None
            vid_c = vid_col_r if vid_col_r != na else None

            work_df = rdf.head(max_records) if max_records > 0 else rdf
            total = len(work_df)

            # ── Phase 1 ─────────────────────────────────────────────────
            st.markdown("---")

            # Determine URL column availability
            url_c_col = url_col if url_col != na else None

            # In candidate mode, split records into URL vs no-URL groups
            if candidate_mode and url_c_col:
                has_url_mask = work_df[url_c_col].apply(
                    lambda x: bool(x) and pd.notna(x) and str(x).strip() not in ('', 'nan', 'None'))
                p1_indices = work_df[has_url_mask].index.tolist()
                skip_p1_indices = work_df[~has_url_mask].index.tolist()
            elif candidate_mode and not url_c_col:
                # No URL column mapped at all — everything skips Phase 1
                p1_indices = []
                skip_p1_indices = work_df.index.tolist()
            else:
                # Standard mode: everything goes through Phase 1
                p1_indices = work_df.index.tolist()
                skip_p1_indices = []

            phase1 = {}

            # Handle records that skip Phase 1 (candidate mode, no URL)
            if skip_p1_indices:
                st.markdown(f"#### \U0001f7e3 Candidate Records (No URL)")
                st.markdown(
                    f"<p style='color:#a78bfa;font-size:0.9rem;'>"
                    f"<strong>{len(skip_p1_indices):,}</strong> candidate records have no URL "
                    f"&mdash; skipping Phase 1, will go directly to Phase 2 web search.</p>",
                    unsafe_allow_html=True)
                for idx in skip_p1_indices:
                    phase1[idx] = dict(
                        url_ok=False, hcp_score=0, hco_score=0, dept_score=0,
                        status='skipped_candidate', notes='Candidate record \u2014 no URL, sent to Phase 2')

            # Phase 1 for records WITH URLs
            if p1_indices:
                st.markdown("#### \U0001f7e2 Phase 1 \u2014 Checking Existing URLs")
                p1_bar = st.progress(0)
                p1_txt = st.empty()
                p1_stats = st.empty()

                cnt = dict(confirmed=0, partial=0, failed=0)
                p1_total = len(p1_indices)

                def _do_p1(idx_row):
                    idx, row = idx_row
                    return idx, _check_url(
                        row.get(url_c_col, '') if url_c_col else '',
                        str(row.get(fn_col, '')),
                        str(row.get(ln_col, '')),
                        str(row.get(hco_col, '')),
                        str(row.get(dept_c, '')) if dept_c else '',
                        str(row.get(spec_c, '')) if spec_c else '')

                with ThreadPoolExecutor(max_workers=max_workers) as ex:
                    futs = {ex.submit(_do_p1, (i, work_df.loc[i])): i for i in p1_indices}
                    done = 0
                    for f in as_completed(futs):
                        idx, res = f.result()
                        phase1[idx] = res
                        done += 1
                        if res['status'] == 'confirmed':
                            cnt['confirmed'] += 1
                        elif 'partial' in res['status']:
                            cnt['partial'] += 1
                        else:
                            cnt['failed'] += 1
                        p1_bar.progress(done / p1_total)
                        p1_txt.text(f"Processed {done:,} / {p1_total:,} records")
                        p1_stats.markdown(
                            f"\u2705 Confirmed: **{cnt['confirmed']:,}** &nbsp;\u2022&nbsp; "
                            f"\U0001f7e1 Partial: **{cnt['partial']:,}** &nbsp;\u2022&nbsp; "
                            f"\u274c Failed: **{cnt['failed']:,}**")

                p1_bar.progress(1.0)
                st.success(f"Phase 1 complete: {cnt['confirmed']:,} confirmed out of {p1_total:,}")

            # ── Phase 2 ─────────────────────────────────────────────────
            phase2 = {}
            if run_phase2:
                need_p2 = [i for i, r in phase1.items() if r['status'] != 'confirmed']
                if need_p2:
                    candidate_in_p2 = sum(
                        1 for i in need_p2
                        if phase1.get(i, {}).get('status') == 'skipped_candidate')
                    if candidate_in_p2 > 0:
                        st.markdown(
                            f"#### \U0001f535 Phase 2 \u2014 Web Search "
                            f"({candidate_in_p2:,} candidates"
                            f"{f' + {len(need_p2) - candidate_in_p2:,} failed Phase 1' if len(need_p2) - candidate_in_p2 > 0 else ''})")
                    else:
                        st.markdown("#### \U0001f535 Phase 2 \u2014 Web Search for Unconfirmed Records")
                    p2_bar = st.progress(0)
                    p2_txt = st.empty()
                    p2_stats = st.empty()
                    p2_found = 0
                    p2_total = len(need_p2)

                    def _do_p2(idx):
                        row = work_df.loc[idx]
                        return idx, _web_search(
                            str(row.get(fn_col, '')),
                            str(row.get(ln_col, '')),
                            str(row.get(hco_col, '')),
                            str(row.get(spec_c, '')) if spec_c else '',
                            str(row.get(city_c, '')) if city_c else '')

                    with ThreadPoolExecutor(max_workers=min(max_workers, 5)) as ex:
                        futs = {ex.submit(_do_p2, idx): idx for idx in need_p2}
                        done2 = 0
                        p2_suggested = 0
                        for f in as_completed(futs):
                            idx, res = f.result()
                            phase2[idx] = res
                            done2 += 1
                            if res['found']:
                                p2_found += 1
                            elif res.get('suggested_url'):
                                p2_suggested += 1
                            p2_bar.progress(done2 / p2_total)
                            p2_txt.text(f"Searched {done2:,} / {p2_total:,} records")
                            p2_stats.markdown(
                                f"\U0001f50d Searched: **{done2:,}** &nbsp;\u2022&nbsp; "
                                f"\u2705 Confirmed: **{p2_found:,}** &nbsp;\u2022&nbsp; "
                                f"\U0001f4a1 Leads: **{p2_suggested:,}**")

                    p2_bar.progress(1.0)
                    st.success(f"Phase 2 complete: {p2_found:,} confirmed + "
                               f"{p2_suggested:,} suggested leads for manual review")

                    # ── Phase 2 Diagnostics ───────────────────────────────
                    with st.expander("🔍 Phase 2 Search Diagnostics (click to expand)"):
                        err_count = 0
                        for idx, res in phase2.items():
                            dbg = res.get('search_debug', '')
                            if dbg and ('error' in dbg.lower() or 'failed' in dbg.lower()):
                                err_count += 1
                        if err_count > 0:
                            st.warning(
                                f"⚠️ {err_count} out of {len(phase2)} searches had backend "
                                f"errors. This usually means DuckDuckGo/Google is rate-limiting "
                                f"or blocking requests. Try reducing concurrency or adding delays.")
                        else:
                            st.info(
                                f"✅ All {len(phase2)} searches completed without backend errors.")
                        # Show sample debug for first few unconfirmed
                        samples = [(i, r) for i, r in phase2.items()
                                   if not r.get('found') and r.get('search_debug')][:5]
                        if samples:
                            st.markdown("**Sample search logs (first 5 unconfirmed):**")
                            for idx, res in samples:
                                row_data = work_df.loc[idx]
                                name = f"{row_data.get(fn_col, '')} {row_data.get(ln_col, '')}"
                                st.code(f"{name.strip()}\n{res.get('search_debug', 'No debug info')}",
                                        language='text')

            # ── Build Results DataFrame ──────────────────────────────────
            rows = []
            for i, row in work_df.iterrows():
                p1 = phase1.get(i, {})
                p2 = phase2.get(i, {})
                suggested = p2.get('suggested_url', '') if p2 else ''
                confidence = p2.get('confidence', '') if p2 else ''
                aff_confirmed = False
                validation_url = ''

                _orig_url = str(row.get(url_c_col, '')) if url_c_col else ''

                if p1.get('status') == 'confirmed':
                    vs = 'Confirmed'
                    src = _orig_url
                    nurl = ''
                    notes = p1.get('notes', '')
                    aff_confirmed = True
                    validation_url = src
                elif p2.get('found'):
                    vs = 'Confirmed (New URL)'
                    src = p2.get('source', '')
                    nurl = p2.get('new_url', '')
                    notes = p2.get('notes', '')
                    confidence = p2.get('confidence', 'High')
                    aff_confirmed = p2.get('affiliation_confirmed', False)
                    validation_url = nurl
                elif suggested:
                    # Has a suggested URL but not strong enough to auto-confirm
                    vs = 'Suggested Lead'
                    src = p2.get('source', '')
                    nurl = ''
                    notes = p2.get('notes', '')
                    aff_confirmed = False
                    validation_url = suggested
                elif 'partial' in p1.get('status', ''):
                    vs = 'Partial Match'
                    src = _orig_url
                    nurl = ''
                    notes = p1.get('notes', '')
                    aff_confirmed = False
                    validation_url = src
                else:
                    vs = 'Unconfirmed'
                    src = ''
                    nurl = ''
                    notes = p1.get('notes', '')
                    if p2:
                        notes += ' | ' + p2.get('notes', '')

                r = {}
                if vid_c:
                    r['VID'] = row.get(vid_c, '')
                r['First_Name'] = row.get(fn_col, '')
                r['Last_Name'] = row.get(ln_col, '')
                r['HCO_Name'] = row.get(hco_col, '')
                if dept_c:
                    r['Department'] = row.get(dept_c, '')
                r['Record_Type'] = 'Candidate' if p1.get('status') == 'skipped_candidate' else 'Existing'
                r['Affiliation_Confirmed'] = 'Yes' if aff_confirmed else 'No'
                r['Validation_Status'] = vs
                r['Validation_URL'] = validation_url
                r['HCP_Score'] = p1.get('hcp_score', 0)
                r['HCO_Score'] = p1.get('hco_score', 0)
                r['Original_URL'] = _orig_url
                r['New_URL'] = nurl
                r['Suggested_URL'] = suggested
                r['Confidence'] = confidence
                r['Source'] = src
                r['Notes'] = notes
                r['Search_Debug'] = p2.get('search_debug', '') if p2 else ''
                rows.append(r)

            df_results = pd.DataFrame(rows)
            st.session_state.reval_results = df_results
            st.rerun()

    # ── Display Results ──────────────────────────────────────────────────
    if 'reval_results' in st.session_state and st.session_state.reval_results is not None:
        dfr = st.session_state.reval_results
        st.markdown("---")
        st.markdown("### Revalidation Results")

        total_r = len(dfr)
        conf_r = len(dfr[dfr['Validation_Status'] == 'Confirmed'])
        new_r = len(dfr[dfr['Validation_Status'] == 'Confirmed (New URL)'])
        lead_r = len(dfr[dfr['Validation_Status'] == 'Suggested Lead'])
        part_r = len(dfr[dfr['Validation_Status'] == 'Partial Match'])
        unconf_r = len(dfr[dfr['Validation_Status'] == 'Unconfirmed'])
        auto_rate = (conf_r + new_r) / max(total_r, 1) * 100
        total_with_urls = conf_r + new_r + lead_r
        coverage_rate = total_with_urls / max(total_r, 1) * 100

        # Affiliation confirmation counts
        aff_yes = len(dfr[dfr['Affiliation_Confirmed'] == 'Yes']) if 'Affiliation_Confirmed' in dfr.columns else 0
        aff_no = total_r - aff_yes
        aff_rate = aff_yes / max(total_r, 1) * 100

        # Primary goal metric: Affiliation Confirmation
        aff_color = '#34d399' if aff_rate >= 60 else '#fbbf24' if aff_rate >= 40 else '#f87171'
        st.markdown(f"""
        <div class="reval-rate-box" style="border:2px solid {aff_color};">
            <p class="label">HCP-HCO Affiliation Confirmation Rate</p>
            <p class="rate" style="color:{aff_color};">{aff_rate:.1f}%</p>
            <p class="sub">{aff_yes:,} affiliations confirmed via URL &nbsp;\u2022&nbsp;
                {aff_no:,} unconfirmed (need manual review or new sources)</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        mc1, mc2, mc3, mc4, mc5, mc6, mc7 = st.columns(7)
        mc1.metric("Total", f"{total_r:,}")
        mc2.metric("Affil. Confirmed", f"{aff_yes:,}")
        mc3.metric("Confirmed (URL)", f"{conf_r:,}")
        mc4.metric("Confirmed (New)", f"{new_r:,}")
        mc5.metric("Suggested Leads", f"{lead_r:,}", help="URLs found but need manual verification")
        mc6.metric("Partial Match", f"{part_r:,}")
        mc7.metric("No URL Found", f"{unconf_r:,}")

        # Show candidate vs existing breakdown if applicable
        if 'Record_Type' in dfr.columns:
            cand_count = len(dfr[dfr['Record_Type'] == 'Candidate'])
            exist_count = len(dfr[dfr['Record_Type'] == 'Existing'])
            if cand_count > 0:
                st.markdown("<br>", unsafe_allow_html=True)
                cc1, cc2, cc3 = st.columns(3)
                cc1.metric("\U0001f7e3 Candidate Records", f"{cand_count:,}",
                           help="Records without Source URL (skipped Phase 1)")
                cc2.metric("\U0001f7e2 Existing Records", f"{exist_count:,}",
                           help="Records with Source URL (went through Phase 1)")
                cand_confirmed = len(dfr[(dfr['Record_Type'] == 'Candidate') &
                                         (dfr['Affiliation_Confirmed'] == 'Yes')])
                cand_rate = cand_confirmed / max(cand_count, 1) * 100
                cand_color = '#34d399' if cand_rate >= 60 else '#fbbf24' if cand_rate >= 40 else '#f87171'
                cc3.metric("\U0001f50d Candidate Confirmation Rate", f"{cand_rate:.1f}%",
                           help="How many candidate records got affiliation confirmed via web search")

        st.markdown("<br>", unsafe_allow_html=True)

        rate_color = '#34d399' if auto_rate >= 60 else '#fbbf24' if auto_rate >= 40 else '#f87171'
        cov_color = '#34d399' if coverage_rate >= 75 else '#fbbf24' if coverage_rate >= 50 else '#f87171'
        st.markdown(f"""
        <div class="reval-rate-box">
            <p class="label">Automation Rate (Auto-Confirmed)</p>
            <p class="rate" style="color:{rate_color};">{auto_rate:.1f}%</p>
            <p class="sub">{conf_r + new_r:,} auto-confirmed &nbsp;\u2022&nbsp;
                {lead_r:,} suggested leads for manual review</p>
        </div>
        <div class="reval-rate-box" style="margin-top:12px;">
            <p class="label">URL Coverage (Confirmed + Suggested Leads)</p>
            <p class="rate" style="color:{cov_color};">{coverage_rate:.1f}%</p>
            <p class="sub">{total_with_urls:,} records have a URL &nbsp;\u2022&nbsp;
                {unconf_r + part_r:,} truly need manual search</p>
        </div>
        """, unsafe_allow_html=True)

        ch1, ch2, ch3 = st.columns(3)
        with ch1:
            fig_aff = make_donut(aff_yes, aff_no,
                                 'Affiliation Confirmed', 'Unconfirmed',
                                 '#34d399', '#f87171', 'HCP-HCO Affiliation')
            if fig_aff:
                st.plotly_chart(fig_aff, use_container_width=True)
        with ch2:
            fig = make_donut(conf_r + new_r, lead_r + unconf_r + part_r,
                             'Auto-Confirmed', 'Needs Review',
                             '#60a5fa', '#fbbf24', 'Revalidation Outcome')
            if fig:
                st.plotly_chart(fig, use_container_width=True)
        with ch3:
            fig2 = go.Figure(data=[go.Bar(
                x=['Confirmed\n(URL)', 'New URL\nFound', 'Suggested\nLeads',
                   'Partial\nMatch', 'No URL\nFound'],
                y=[conf_r, new_r, lead_r, part_r, unconf_r],
                marker_color=['#34d399', '#60a5fa', '#a78bfa', '#fbbf24', '#f87171'],
                text=[conf_r, new_r, lead_r, part_r, unconf_r],
                textposition='outside',
            )])
            fig2.update_layout(
                title=dict(text='Status Breakdown', font=dict(size=15, family='Inter')),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                height=300, margin=dict(t=50, b=20, l=10, r=10),
                yaxis=dict(showgrid=False, showticklabels=False),
                xaxis=dict(tickfont=dict(size=11)),
            )
            st.plotly_chart(fig2, use_container_width=True)

        st.markdown("### Detailed Results")
        st.dataframe(dfr, use_container_width=True, height=500)

        reval_excel = export_reval_excel(
            dfr, st.session_state.get('reval_file_name', 'revalidation'))
        st.download_button(
            "\U0001f4e5 Download Revalidation Report",
            data=reval_excel,
            file_name=f"Revalidation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
